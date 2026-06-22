#!/usr/bin/env python3
"""
创建 Excel 导入模板
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_excel_template():
    """创建 Excel 导入模板文件"""
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "录取数据导入模板"
    
    # 定义字段配置 (字段名，说明，是否必填，示例值)
    fields = [
        ("student_name", "学生姓名", True, "张三"),
        ("student_name_en", "学生英文名", False, "San Zhang"),
        ("source_school", "来源学校", False, "北京四中"),
        ("country", "国家/地区", True, "美国"),
        ("university", "录取大学", True, "University of California, Berkeley"),
        ("university_en", "录取大学 (中文)", False, "加州大学伯克利分校"),
        ("major", "录取专业", False, "Computer Science"),
        ("major_en", "专业 (中文)", False, "计算机科学"),
        ("degree", "学位", False, "本科"),
        ("gpa", "GPA", False, "3.85"),
        ("toefl", "TOEFL 分数", False, "105"),
        ("ielts", "IELTS 分数", False, "7.5"),
        ("sat", "SAT 分数", False, "1520"),
        ("act", "ACT 分数", False, "34"),
        ("gre", "GRE 分数", False, "325"),
        ("gmat", "GMAT 分数", False, "720"),
        ("scholarship", "奖学金名称", False, "Merit Scholarship"),
        ("scholarship_amount", "奖学金金额 (USD)", False, "5000"),
        ("application_year", "申请年份", False, "2026"),
        ("admission_year", "入学年份", False, "2026"),
        ("article_url", "文章来源 URL", False, "https://mp.weixin.qq.com/s/xxx"),
        ("article_title", "文章标题", False, "2026 届录取喜报"),
        ("requirements", "录取条件", False, "GPA 3.5+, TOEFL 100+"),
        ("portfolio", "作品集要求", False, "需要提交作品集"),
        ("is_verified", "是否已验证", False, "1"),
    ]
    
    # 设置表头样式
    header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头 (第 1 行)
    headers = ["字段名 (英文)", "字段说明", "是否必填", "示例值", "填写说明"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 填写说明
    instructions = [
        "请勿修改此列，用于识别字段",
        "字段的中文说明",
        "是/否，必填字段必须填写",
        "参考填写格式",
        "详细的填写要求和注意事项"
    ]
    
    # 写入字段说明 (第 2 行开始)
    detailed_instructions = [
        "学生中文姓名，必填",
        "学生英文姓名，拼音格式",
        "学生毕业的高中/学校名称",
        "国家或地区名称，如：美国、英国、加拿大",
        "录取大学的英文名称，必填",
        "录取大学的中文名称",
        "录取专业的英文名称",
        "录取专业的中文名称",
        "学位类型：本科/硕士/博士",
        "GPA 成绩，范围 0-4.0",
        "TOEFL 总分，范围 0-120",
        "IELTS 总分，范围 0-9.0",
        "SAT 总分，范围 400-1600",
        "ACT 总分，范围 1-36",
        "GRE 总分，范围 260-340",
        "GMAT 总分，范围 200-800",
        "奖学金名称或类型",
        "奖学金金额，美元为单位",
        "申请季年份，如 2026",
        "实际入学年份，如 2026",
        "录取信息来源的文章链接",
        "录取信息来源的文章标题",
        "录取的学术条件要求",
        "是否需要作品集及说明",
        "是否已验证：1=已验证，0=待验证"
    ]
    
    for row, (field, desc, required, example) in enumerate(fields, 2):
        # 字段名
        cell = ws.cell(row=row, column=1, value=field)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # 字段说明
        cell = ws.cell(row=row, column=2, value=desc)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # 是否必填
        required_text = "是" if required else "否"
        cell = ws.cell(row=row, column=3, value=required_text)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if required:
            cell.font = Font(bold=True, color="FF0000")
        
        # 示例值
        cell = ws.cell(row=row, column=4, value=example)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.font = Font(italic=True, color="666666")
        
        # 填写说明
        cell = ws.cell(row=row, column=5, value=detailed_instructions[row-2])
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # 设置列宽
    column_widths = [20, 25, 12, 35, 40]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 冻结首行
    ws.freeze_panes = "A2"
    
    # 创建数据填写示例工作表
    ws_sample = wb.create_sheet(title="填写示例")
    
    # 示例数据表头
    sample_headers = [f[0] for f in fields]
    sample_data = [
        ["张三", "San Zhang", "北京四中", "美国", "University of California, Berkeley", "加州大学伯克利分校", 
         "Computer Science", "计算机科学", "本科", "3.85", "105", "", "1520", "", "", "", 
         "Merit Scholarship", "5000", "2026", "2026", "https://example.com/article1", "2026 届录取喜报", 
         "GPA 3.5+, TOEFL 100+", "不需要", "1"],
        ["李四", "Si Li", "上海中学", "美国", "Stanford University", "斯坦福大学",
         "Economics", "经济学", "本科", "3.92", "110", "8.0", "", "", "330", "",
         "Full Ride", "75000", "2026", "2026", "https://example.com/article2", "斯坦福录取",
         "GPA 3.8+, SAT 1500+", "需要提交文书", "1"],
        ["王五", "Wu Wang", "深圳中学", "英国", "University of Oxford", "牛津大学",
         "Mathematics", "数学", "本科", "", "", "7.5", "", "", "", "",
         "Clarendon Fund", "", "2026", "2026", "https://example.com/article3", "牛津数学录取",
         "A-Level A*AA", "不需要", "0"],
    ]
    
    # 写入示例数据表头
    sample_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    sample_font = Font(bold=True, color="FFFFFF", size=11)
    sample_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(sample_headers, 1):
        cell = ws_sample.cell(row=1, column=col, value=header)
        cell.fill = sample_fill
        cell.font = sample_font
        cell.alignment = sample_alignment
        cell.border = thin_border
    
    # 写入示例数据
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_sample.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # 设置示例表列宽
    for col in range(1, len(sample_headers) + 1):
        ws_sample.column_dimensions[get_column_letter(col)].width = 18
    
    ws_sample.freeze_panes = "A2"
    
    # 创建说明工作表
    ws_guide = wb.create_sheet(title="导入说明")
    
    guide_content = [
        ["📋 Excel 导入模板使用说明", ""],
        ["", ""],
        ["1️⃣ 下载模板", "从系统下载 Excel 导入模板文件"],
        ["", ""],
        ["2️⃣ 填写数据", "在「填写示例」工作表中复制格式，填写您的录取数据"],
        ["   - 必填字段：学生姓名、国家/地区、录取大学", ""],
        ["   - 数字字段：GPA、标化成绩等填写数字即可，不要带单位", ""],
        ["   - 是/否字段：is_verified 填写 1(已验证) 或 0(待验证)", ""],
        ["", ""],
        ["3️⃣ 数据验证", "系统会自动验证以下项目：", ""],
        ["   ✓ 必填字段检查", "确保必填字段不为空"],
        ["   ✓ 格式检查", "数字字段必须是有效数字"],
        ["   ✓ 范围检查", "成绩分数在合理范围内"],
        ["   ✓ 重复检测", "检测学生姓名 + 大学 + 入学年份是否重复"],
        ["", ""],
        ["4️⃣ 上传导入", "将填写好的 Excel 文件上传到系统，等待导入完成"],
        ["", ""],
        ["5️⃣ 查看结果", "系统会显示导入结果：成功数量、跳过数量、错误详情"],
        ["", ""],
        ["⚠️ 注意事项", ""],
        ["   - 单次最多支持导入 1000 条记录", ""],
        ["   - 重复数据会自动跳过（相同学生 + 大学 + 入学年份）", ""],
        ["   - 错误数据会返回详细错误信息，修正后可重新导入", ""],
        ["   - 建议先导入少量数据测试，确认格式正确后再批量导入", ""],
        ["", ""],
        ["📞 如有问题，请联系系统管理员", ""],
    ]
    
    for row_idx, row_data in enumerate(guide_content, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_guide.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
    
    ws_guide.column_dimensions['A'].width = 40
    ws_guide.column_dimensions['B'].width = 60
    
    # 保存文件
    output_path = "/root/.openclaw/workspace/大学录取信息整理系统/backend/app/templates/录取数据导入模板.xlsx"
    wb.save(output_path)
    
    return output_path

if __name__ == "__main__":
    import os
    os.makedirs("/root/.openclaw/workspace/大学录取信息整理系统/backend/app/templates", exist_ok=True)
    output_path = create_excel_template()
    print(f"✅ Excel 模板已创建：{output_path}")
