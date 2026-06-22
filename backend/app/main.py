# 完成时间：2026-03-19 02:45 UTC - 任务队列管理界面已完成
"""
大学录取信息整理系统 - FastAPI 主应用
单端口部署方案：前后端合并，FastAPI 服务静态文件和 API
"""
from fastapi import FastAPI, HTTPException, Query, Request, UploadFile, File, Form, Depends
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse, HTMLResponse, StreamingResponse, Response
from fastapi.middleware.cors import CORSMiddleware
import os
import sqlite3
import pymysql
import io
import time
import logging
import json
import re
from datetime import datetime, timedelta
from typing import Dict, Any, List
import pandas as pd
from openpyxl import load_workbook
from openai import OpenAI
from app.db_config import get_db_connection, exec_sql, DB_TYPE
from app.admin_page import HTML_CONTENT as ADMIN_HTML

# 导入 PDF 模板（延迟导入，避免缺少 reportlab 时无法启动）
# from app.pdf_template import create_pdf_report

# 导入监控模块（Agent-11 新增）
from app.monitor import monitor_service

# 导入录取信息状态查询模块（2026-04-07 新增）
from api.admission_status import router as admission_status_router

# 导入用户认证 API 模块
from app.auth_api import router as auth_api_router

# 导入用户认证模块（Agent-10 新增）
from app.auth import (
    authenticate_user, create_access_token, get_current_user, has_permission,
    create_user, get_all_users, get_all_roles, update_user_role, deactivate_user,
    activate_user, get_user_by_username, update_last_login
)
from pydantic import BaseModel

# 数据目录
DATA_DIR = os.path.join(os.path.dirname(__file__), "../data")
os.makedirs(DATA_DIR, exist_ok=True)

# 创建 FastAPI 应用
app = FastAPI(title="大学录取信息整理系统", version="1.0.0")

# CORS
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

# 注册录取信息状态查询路由（2026-04-07 新增）
app.include_router(admission_status_router, prefix="/api")

# 注册用户认证 API 路由
app.include_router(auth_api_router)

# API 响应时间监控中间件（Agent-11 新增）
@app.middleware("http")
async def monitor_api_latency(request: Request, call_next):
    start_time = time.time()
    response = await call_next(request)
    process_time = time.time() - start_time
    
    # 记录 API 响应时间（只记录/api/开头的接口）
    if request.url.path.startswith("/api/"):
        monitor_service.record_api_latency(
            endpoint=request.url.path,
            method=request.method,
            latency_sec=process_time,
            status_code=response.status_code
        )
    
    return response

def get_local_time():
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')

def get_db_connection():
    from app.db_config import get_db_connection as _get_db
    return _get_db()

def get_fetch_value(result):
    """兼容 MySQL DictCursor 和 SQLite Row 的值提取"""
    if isinstance(result, dict):
        return result.get('COUNT(*)', result.get('COUNT(DISTINCT country)', result.get('COUNT(DISTINCT university_cn)', list(result.values())[0] if result else 0)))
    else:
        return result[0] if result else 0

def get_pooled_connection():
    from app.task_queue import get_pooled_connection as _pooled
    return _pooled()

def init_db():
    """初始化数据库 - 兼容 MySQL 和 SQLite
    """
    from app.db_config import DB_TYPE, exec_sql
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    exec_sql(cursor, "SELECT 1")
    
    if DB_TYPE == "mysql":
        cursor.execute("""CREATE TABLE IF NOT EXISTS concurrency_settings (
            id INT PRIMARY KEY,
            enabled TINYINT DEFAULT 0,
            max_concurrent INT DEFAULT 1,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
        )""")
        cursor.execute("INSERT IGNORE INTO concurrency_settings (id, enabled, max_concurrent) VALUES (1, 0, 2)")
    else:
        cursor.execute("""CREATE TABLE IF NOT EXISTS concurrency_settings (
            id INTEGER PRIMARY KEY CHECK (id = 1),
            enabled INTEGER DEFAULT 0,
            max_concurrent INTEGER DEFAULT 1,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""")
        cursor.execute("INSERT OR IGNORE INTO concurrency_settings (id, enabled, max_concurrent) VALUES (1, 0, 2)")
    
    conn.commit()
    conn.close()
    print(f"[OK] 数据库连接正常 ({DB_TYPE})")
    print("[OK] concurrency_settings 表已就绪")

    _recover_stuck_tasks_on_startup()

def _recover_stuck_tasks_on_startup():
    """服务启动时恢复长时间卡在状态2（处理中）的任务"""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE collection_tasks SET task_status = 0, started_at = NULL, "
        "error_message = '服务重启恢复：任务卡在处理中状态' "
        "WHERE task_status = 2"
    )
    recovered = cursor.rowcount
    conn.commit()
    conn.close()
    if recovered > 0:
        print(f"[RECOVERY] 服务启动时恢复了 {recovered} 个卡住的任务")

# 初始化数据库
init_db()

def write_log(user_id=0, username="system", operation_type="", operation_desc=""):
    """写操作日志辅助函数"""
    import socket
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        exec_sql(cursor, "INSERT INTO operation_logs (user_id, username, operation_type, operation_desc, ip_address) VALUES (?, ?, ?, ?, ?)", [user_id, username, operation_type, operation_desc, ""])
        conn.commit()
        conn.close()
    except:
        pass



@app.get("/", response_class=HTMLResponse)
async def root():
    """管理后台首页"""
    return HTMLResponse(ADMIN_HTML)


@app.get("/api/records")
async def get_records(page: int = 1, page_size: int = 50, search: str = "", country: str = "", year: str = ""):
    """简化版 API - 支持搜索和筛选"""
    from app.db_config import exec_sql, DB_TYPE
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 构建查询条件
    where_clauses = []
    params = []
    
    if search:
        where_clauses.append("(student_name_cn LIKE ? OR university_cn LIKE ? OR major_cn LIKE ? OR source_school LIKE ?)")
        search_pattern = f"%{search}%"
        params.extend([search_pattern, search_pattern, search_pattern, search_pattern])
    
    if country:
        where_clauses.append("country = ?")
        params.append(country)
    
    if year:
        where_clauses.append("admission_year = ?")
        params.append(year)
    
    where_sql = " WHERE " + " AND ".join(where_clauses) if where_clauses else ""
    
    # 查询记录
    query = f"SELECT * FROM admission_records{where_sql} ORDER BY created_at DESC LIMIT ? OFFSET ?"
    exec_sql(cursor, query, params + [page_size, (page - 1) * page_size])
    rows = cursor.fetchall()
    
    # 查询总数
    count_query = f"SELECT COUNT(*) FROM admission_records{where_sql}"
    exec_sql(cursor, count_query, params)
    total = get_fetch_value(cursor.fetchone())
    conn.close()
    
    # 转换字段名以兼容前端（驼峰命名）
    records = []
    for row in rows:
        record = dict(row)
        # 添加前端期望的字段名（驼峰命名）
        record['studentName'] = record.get('student_name_cn', '') or record.get('student_name', '')
        record['university'] = record.get('university_cn', '') or record.get('university', '')
        record['major'] = record.get('major_cn', '') or record.get('major', '')
        record['year'] = record.get('admission_year', '')
        record['degree'] = record.get('admission_type', '')
        record['scholarship'] = record.get('scholarship_amount', '')
        record['status'] = record.get('admission_status', '已录取')
        records.append(record)
    
    return {"records": records, "total": total, "page": page, "page_size": page_size, "total_pages": (total + page_size - 1) // page_size}


# ==================== 批量操作 API（Agent-13 新增） ====================

@app.put("/api/records/batch")
async def batch_update_records(request: Request, current_user: dict = Depends(get_current_user)):
    """
    批量修改录取记录字段
    
    请求格式：
    {
        "record_ids": [1, 2, 3, ...],
        "fields": {
            "country_id": 1,
            "university_id": 5,
            "major": "Computer Science",
            "status": "verified",
            "is_verified": 1,
            "degree": "本科",
            "scholarship": "Merit",
            "scholarship_amount": 5000
        }
    }
    
    支持批量修改的字段：
    - country_id: 国家 ID
    - university_id: 大学 ID
    - major: 专业
    - major_en: 专业英文
    - degree: 学位
    - scholarship: 奖学金
    - scholarship_amount: 奖学金金额
    - is_verified: 验证状态
    - status: 状态文本
    - requirements: 录取条件
    - portfolio: 作品集
    - article_url: 文章链接
    """
    data = await request.json()
    record_ids = data.get("record_ids", [])
    fields = data.get("fields", {})
    
    if not record_ids:
        raise HTTPException(status_code=400, detail="请提供要修改的记录 ID 列表")
    
    if not fields:
        raise HTTPException(status_code=400, detail="请提供要修改的字段")
    
    # 允许批量修改的字段
    allowed_fields = [
        "country", "university", "university_en", "major", "major_en", "degree",
        "scholarship", "scholarship_amount", "scholarship_currency", "is_verified",
        "requirements", "portfolio", "article_url", "article_title",
        "gpa", "toefl", "ielts", "sat", "act", "gre", "gmat",
        "student_grade", "country_en", "university_type", "university_ranking",
        "major_category", "admission_type", "admission_status", "conditional_offer",
        "admission_date", "language_requirement_type", "language_score_required",
        "language_score_actual", "language_waived", "sat_required", "test_optional",
        "publish_date", "data_quality", "notes"
    ]
    
    # 过滤只保留允许的字段
    update_fields = {k: v for k, v in fields.items() if k in allowed_fields}
    
    if not update_fields:
        raise HTTPException(status_code=400, detail="没有可更新的字段")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # 1. 备份原始数据（用于撤销）
        placeholders = ','.join('?' * len(record_ids))
        cursor.execute(f"SELECT * FROM admission_records WHERE id IN ({placeholders})", record_ids)
        original_records = [dict(row) for row in cursor.fetchall()]
        
        if not original_records:
            conn.close()
            raise HTTPException(status_code=404, detail="未找到任何记录")
        
        # 2. 创建撤销记录
        import json
        undo_data = {
            "operation": "batch_update",
            "record_ids": record_ids,
            "original_data": original_records,
            "updated_fields": update_fields,
            "timestamp": datetime.now().isoformat(),
            "user_id": current_user.get("id")
        }
        
        cursor.execute("""
            INSERT INTO batch_operations (operation_type, operation_data, created_by, created_at)
            VALUES (?, ?, ?, CURRENT_TIMESTAMP)
        """, ["batch_update", json.dumps(undo_data, ensure_ascii=False), current_user.get("id")])
        operation_id = cursor.lastrowid
        
        # 3. 构建更新 SQL
        update_clauses = []
        values = []
        for field, value in update_fields.items():
            update_clauses.append(f"{field} = ?")
            values.append(value)
        update_clauses.append("updated_at = CURRENT_TIMESTAMP")
        
        # 4. 执行批量更新
        values.extend(record_ids)
        query = f"UPDATE admission_records SET {', '.join(update_clauses)} WHERE id IN ({placeholders})"
        cursor.execute(query, values)
        updated_count = cursor.rowcount
        
        conn.commit()
        
        # 5. 返回结果
        return {
            "message": "批量更新成功",
            "updated_count": updated_count,
            "operation_id": operation_id,
            "can_undo": True
        }
        
    except Exception as e:
        conn.rollback()
        conn.close()
        raise HTTPException(status_code=500, detail=f"批量更新失败：{str(e)}")



@app.post("/api/verify-password")
async def verify_password(request: Request):
    """验证用户密码（用于批量删除等敏感操作）"""
    from app.auth import authenticate_user
    data = await request.json()
    username = data.get("username", "")
    password = data.get("password", "")
    
    if not username or not password:
        raise HTTPException(status_code=400, detail="请输入账号和密码")
    
    user = authenticate_user(username, password)
    if not user:
        raise HTTPException(status_code=401, detail="账号或密码错误")
    
    return {"success": True, "message": "验证成功"}


@app.delete("/api/records/batch")
async def batch_delete_records(request: Request):
    """
    批量删除录取记录
    
    请求格式：
    {
        "record_ids": [1, 2, 3, ...],
        "confirm": true  // 需要确认标志
    }
    """
    data = await request.json()
    record_ids = data.get("record_ids", [])
    confirm = data.get("confirm", False)
    
    if not record_ids:
        raise HTTPException(status_code=400, detail="请提供要删除的记录 ID 列表")
    
    if not confirm:
        raise HTTPException(status_code=400, detail="需要确认删除操作")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # 临时关闭外键检查（review_history 等表有 FK 引用）
        exec_sql(cursor, "PRAGMA foreign_keys = OFF")
        
        # 执行批量删除
        placeholders = ','.join('?' * len(record_ids))
        cursor.execute(f"SELECT id FROM admission_records WHERE id IN ({placeholders})", record_ids)
        found_ids = [row[0] for row in cursor.fetchall()]
        
        if not found_ids:
            conn.close()
            raise HTTPException(status_code=404, detail="未找到任何记录")
        
        # 执行删除
        cursor.execute(f"DELETE FROM admission_records WHERE id IN ({placeholders})", record_ids)
        deleted_count = cursor.rowcount
        
        exec_sql(cursor, "PRAGMA foreign_keys = ON")
        
        conn.commit()
        conn.close()
        
        write_log(operation_type="batch_delete", operation_desc=f"批量删除 {deleted_count} 条录取记录")
        
        return {
            "message": "批量删除成功",
            "deleted_count": deleted_count,
        }
        
    except HTTPException:
        conn.close()
        raise
    except Exception as e:
        conn.rollback()
        conn.close()
        raise HTTPException(status_code=500, detail=f"批量删除失败：{str(e)}")


@app.get("/api/records/export")
async def export_records(format: str = Query("excel", description="导出格式：excel 或 csv")):
    """
    导出录取记录为 Excel 或 CSV 格式
    导出字段：学生姓名（中文 + 英文）、来源学校、国家/地区、录取大学（中文 + 英文）、录取专业、奖学金信息、录取条件、状态
    """
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 获取所有记录（使用实际的字段名）
    query = """
    SELECT 
        student_name_cn,
        student_name_en,
        source_school,
        country,
        country_en,
        university_cn,
        university_en,
        major_cn,
        major_en,
        scholarship_amount,
        scholarship_currency,
        scholarship_type,
        language_requirement_type,
        language_score_required,
        language_score_actual,
        sat_required,
        sat_actual,
        admission_status,
        review_status,
        admission_year
    FROM admission_records
    ORDER BY created_at DESC
    """
    
    cursor.execute(query)
    rows = cursor.fetchall()
    conn.close()
    
    # 准备数据
    data = []
    for row in rows:
        # 转换为字典
        row_dict = dict(row)
        
        # 构建录取条件字符串
        requirements = []
        if row_dict.get('language_requirement_type') and row_dict.get('language_score_required'):
            requirements.append(f"{row_dict['language_requirement_type']}: {row_dict['language_score_required']}")
        if row_dict.get('sat_required'):
            requirements.append(f"SAT: {row_dict['sat_required']}")
        
        # 构建奖学金显示
        scholarship_display = ''
        if row_dict.get('scholarship_amount'):
            currency = row_dict.get('scholarship_currency') or '$'
            scholarship_display = f"{currency}{row_dict['scholarship_amount']:,}"
        elif row_dict.get('scholarship_type'):
            scholarship_display = row_dict['scholarship_type']
        
        # 状态
        review_status_map = {
            'pending': '待审核',
            'approved': '已通过',
            'rejected': '已拒绝'
        }
        status = review_status_map.get(row_dict.get('review_status'), row_dict.get('admission_status') or '未知')
        
        data.append({
            '学生姓名': row_dict.get('student_name_cn', ''),
            '学生姓名 (英文)': row_dict.get('student_name_en', ''),
            '来源学校': row_dict.get('source_school', ''),
            '国家/地区': row_dict.get('country', ''),
            '国家/地区 (英文)': row_dict.get('country_en', ''),
            '录取大学': row_dict.get('university_cn', ''),
            '录取大学 (英文)': row_dict.get('university_en', ''),
            '录取专业': row_dict.get('major_cn', ''),
            '录取专业 (英文)': row_dict.get('major_en', ''),
            '奖学金信息': scholarship_display,
            '录取条件': ' | '.join(requirements) if requirements else '',
            '录取状态': row_dict.get('admission_status', ''),
            '审核状态': status,
            '录取年份': row_dict.get('admission_year', '')
        })
    
    if not data:
        raise HTTPException(status_code=404, detail="没有可导出的数据")
    
    # 生成文件名（包含日期）
    date_str = datetime.now().strftime("%Y-%m-%d")
    
    import urllib.parse
    
    if format.lower() == "csv":
        # CSV 导出
        df = pd.DataFrame(data)
        stream = io.StringIO()
        df.to_csv(stream, index=False, encoding='utf-8-sig')
        csv_content = stream.getvalue()
        
        # 使用 RFC 5987 编码处理中文文件名
        filename_encoded = urllib.parse.quote(f"录取数据_{date_str}.csv")
        
        return Response(
            csv_content,
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}"
            }
        )
    else:
        # Excel 导出（默认）
        df = pd.DataFrame(data)
        stream = io.BytesIO()
        with pd.ExcelWriter(stream, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='录取数据')
        
        excel_content = stream.getvalue()
        
        # 使用 RFC 5987 编码处理中文文件名
        filename_encoded = urllib.parse.quote(f"录取数据_{date_str}.xlsx")
        
        return Response(
            excel_content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}"
            }
        )


@app.get("/api/records/{record_id}")
async def get_record(record_id: int):
    """获取单条录取记录详情"""
    conn = get_db_connection()
    cursor = conn.cursor()
    exec_sql(cursor, "SELECT * FROM admission_records WHERE id = ?", (record_id,))
    row = cursor.fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="记录不存在")
    return dict(row)


@app.put("/api/records/{record_id}")
async def update_record(record_id: int, request: Request):
    """更新单条录取记录"""
    data = await request.json()
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 检查记录是否存在
    exec_sql(cursor, "SELECT id FROM admission_records WHERE id = ?", (record_id,))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="记录不存在")
    
    # 构建更新 SQL
    update_clauses = []
    values = []
    for field, value in data.items():
        update_clauses.append(f"{field} = ?")
        values.append(value)
    update_clauses.append("updated_at = CURRENT_TIMESTAMP")
    values.append(record_id)
    
    query = f"UPDATE admission_records SET {', '.join(update_clauses)} WHERE id = ?"
    cursor.execute(query, values)
    conn.commit()
    conn.close()
    write_log(operation_type="update_record", operation_desc=f"更新录取记录 ID={record_id}")
    return {"message": "更新成功"}


@app.delete("/api/records/{record_id}")
async def delete_record(record_id: int):
    """删除单条录取记录"""
    conn = get_db_connection()
    cursor = conn.cursor()
    exec_sql(cursor, "DELETE FROM admission_records WHERE id = ?", (record_id,))
    conn.commit()
    conn.close()
    write_log(operation_type="delete_record", operation_desc=f"删除录取记录 ID={record_id}")
    return {"message": "删除成功"}


@app.get("/api/collection-tasks")
async def get_collection_tasks(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=10, le=5000),
    status: int = Query(None),
    search: str = Query("", max_length=200),
):
    """分段加载任务列表 — 游标分页 + 状态筛选 + 搜索"""
    from app.db_config import DB_TYPE, fetch_count
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    placeholder = "%s" if DB_TYPE == "mysql" else "?"

    where = []
    params = []
    if status is not None:
        where.append(f"task_status = {placeholder}")
        params.append(status)
    if search:
        where.append(f"(article_url LIKE {placeholder} OR source_school_name LIKE {placeholder})")
        p = f"%{search}%"
        params.extend([p, p])

    where_sql = ("WHERE " + " AND ".join(where)) if where else ""

    total = fetch_count(cursor, f"SELECT COUNT(*) FROM collection_tasks {where_sql}", params)

    offset = (page - 1) * page_size
    cursor.execute(
        f"SELECT * FROM collection_tasks {where_sql} ORDER BY priority ASC, created_at DESC LIMIT {placeholder} OFFSET {placeholder}",
        params + [page_size, offset],
    )
    tasks = [dict(row) for row in cursor.fetchall()]
    conn.close()

    total_pages = max(1, (total + page_size - 1) // page_size)
    return {
        "tasks": tasks,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
    }

@app.post("/api/collection-tasks")
async def create_collection_task(request: Request):
    """创建采集任务 — URL 去重 + 大批量批量插入"""
    from app.db_config import DB_TYPE
    import json
    import asyncio as _asyncio
    from app.async_worker import url_seen as _url_seen

    data = await request.json()
    urls = data.get("urls", [])
    if not urls:
        url = data.get("url", "")
        if url:
            urls = [url]
    priority = data.get("priority", 3)
    recognition_model = data.get("recognition_model", "")

    if not urls:
        raise HTTPException(status_code=400, detail="请提供至少一个 URL")

    urls = [url.strip() for url in urls if url and url.strip()]
    if not urls:
        raise HTTPException(status_code=400, detail="请提供至少一个有效的 URL")

    MAX_URLS = 50000
    if len(urls) > MAX_URLS:
        raise HTTPException(status_code=400, detail=f"单次导入 URL 数量不能超过 {MAX_URLS} 条，请分批导入")

    # ── URL 去重 + 已有 URL 过滤 ──
    conn = get_db_connection()
    cursor = conn.cursor()
    exec_sql(cursor, "SELECT article_url FROM collection_tasks")
    existing = {r.get('article_url') if isinstance(r, dict) else r[0] for r in cursor.fetchall() if (r.get('article_url') if isinstance(r, dict) else r[0])}
    conn.close()

    deduped = []
    dup_count = 0
    for url in urls:
        if url in existing:
            dup_count += 1
            continue
        existing.add(url)
        deduped.append(url)

    if not deduped:
        return {"success": True, "created": 0, "duplicates": dup_count, "message": f"所有 {len(urls)} 个 URL 已存在，跳过"}

    BATCH_SIZE = 500
    total_created = 0
    conn = None
    try:
        conn = get_db_connection()
        if DB_TYPE == "sqlite":
            conn.execute("PRAGMA busy_timeout = 60000")
        cursor = conn.cursor()
        if DB_TYPE == "sqlite":
            conn.execute("BEGIN")
        for i in range(0, len(deduped), BATCH_SIZE):
            batch = deduped[i:i + BATCH_SIZE]
            placeholder = "%s" if DB_TYPE == "mysql" else "?"
            cursor.executemany(
                f"INSERT INTO collection_tasks (article_url, title, source_school_name, priority, task_status, recognition_model) VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})",
                [(url, "", "", priority, 0, recognition_model) for url in batch],
            )
            total_created += len(batch)
            conn.commit()
            if i + BATCH_SIZE < len(deduped) and DB_TYPE == "sqlite":
                conn.execute("BEGIN")
        conn.commit()
    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        raise HTTPException(status_code=503, detail=f"数据库繁忙: {e}")
    finally:
        if conn:
            conn.close()

    return {
        "success": True,
        "created": total_created,
        "duplicates": dup_count,
        "message": f"成功创建 {total_created} 个任务，去重 {dup_count} 个",
    }


@app.post("/api/collection/preview")
async def collection_preview(request: Request):
    """采集预览：抓取URL页面，提取标题、学校等基本信息"""
    import urllib.request
    from bs4 import BeautifulSoup
    
    data = await request.json()
    url = data.get("url", "")
    if not url:
        raise HTTPException(status_code=400, detail="请提供 URL")
    
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept": "text/html,application/xhtml+xml",
        }
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=15) as resp:
            html_content = resp.read().decode("utf-8", errors="ignore")
        
        soup = BeautifulSoup(html_content, "html.parser")
        
        # 提取标题
        title = ""
        og_title = soup.find("meta", {"property": "og:title"})
        if og_title:
            title = og_title.get("content", "").strip()
        if not title:
            activity = soup.find(id="activity-name")
            if activity:
                title = activity.get_text(strip=True)
        if not title:
            t = soup.find("title")
            if t:
                title = t.get_text(strip=True)
        
        # 提取学校
        school = ""
        nickname = soup.find("meta", {"property": "og:nickname"})
        if nickname:
            school = nickname.get("content", "").strip()
        if not school:
            ns = soup.find("span", class_="rich_media_meta_nickname")
            if ns:
                school = ns.get_text(strip=True)
        
        # 提取描述
        description = ""
        og_desc = soup.find("meta", {"property": "og:description"})
        if og_desc:
            description = og_desc.get("content", "").strip()
        if not description:
            ds = soup.find("meta", attrs={"name": "description"})
            if ds:
                description = ds.get("content", "").strip()
        
        # 提取封面图
        cover_image = ""
        og_image = soup.find("meta", {"property": "og:image"})
        if og_image:
            cover_image = og_image.get("content", "").strip()
        
        # 估算记录数（表格行数）
        estimated = 0
        tables = soup.find_all("table")
        for table in tables:
            rows = table.find_all("tr")
            if len(rows) > 1:
                estimated += len(rows) - 1
        
        return {
            "success": True,
            "title": title,
            "school": school,
            "description": description,
            "cover_image": cover_image,
            "estimated_records": estimated,
            "message": "预览成功"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        return {"success": False, "message": f"预览失败：{str(e)}"}


@app.post("/api/collection-tasks/{task_id}/start")
async def start_collection_task(task_id: int):
    """开始执行采集任务 — aiohttp + lxml 异步处理"""
    import threading
    import asyncio
    from app.async_worker import process_single_task_async
    from app.db_config import exec_sql, DB_TYPE

    conn = get_db_connection()
    cursor = conn.cursor()
    exec_sql(cursor, "SELECT * FROM collection_tasks WHERE id = ?", [task_id])
    row = cursor.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="任务不存在")
    task = dict(row)
    if task["task_status"] == 2:
        conn.close()
        raise HTTPException(status_code=400, detail="任务已在处理中")
    if task["task_status"] == 3:
        conn.close()
        raise HTTPException(status_code=400, detail="任务已完成")
    exec_sql(cursor, "UPDATE collection_tasks SET task_status = 2, started_at = CURRENT_TIMESTAMP WHERE id = ?", [task_id])
    conn.commit()
    conn.close()

    async def run():
        try:
            await process_single_task_async(task_id)
        except Exception as e:
            import logging
            logger = logging.getLogger('TaskWorker')
            logger.error(f"任务 {task_id} 后台线程异常: {e}", exc_info=True)
            try:
                conn2 = get_db_connection()
                cursor2 = conn2.cursor()
                exec_sql(cursor2, "UPDATE collection_tasks SET task_status = 4, error_message = %s, completed_at = CURRENT_TIMESTAMP WHERE id = %s" if DB_TYPE == "mysql" else "UPDATE collection_tasks SET task_status = 4, error_message = ?, completed_at = CURRENT_TIMESTAMP WHERE id = ?", [str(e)[:500], task_id])
                conn2.commit()
                conn2.close()
            except Exception:
                pass

    threading.Thread(target=lambda: asyncio.run(run()), daemon=True).start()
    return {"message": "任务已启动", "task_id": task_id, "status": "processing"}


@app.post("/api/collection-tasks/batch-start")
async def batch_start_tasks(request: Request):
    """批量启动任务 — aiohttp + lxml 异步并发处理"""
    import threading
    import asyncio
    from app.async_worker import process_single_task_async
    from app.db_config import exec_sql, DB_TYPE

    data = await request.json()
    task_ids = data.get("task_ids", [])

    if not task_ids:
        conn = get_db_connection()
        cursor = conn.cursor()
        exec_sql(cursor, "SELECT id FROM collection_tasks WHERE task_status = 0 ORDER BY priority ASC, created_at ASC LIMIT 1000")
        task_ids = [row[0] if not isinstance(row, dict) else row['id'] for row in cursor.fetchall()]
        conn.close()
        if not task_ids:
            return {"message": "没有待执行的任务", "started": 0}

    MAX_BATCH_SIZE = 500
    if len(task_ids) > MAX_BATCH_SIZE:
        task_ids = task_ids[:MAX_BATCH_SIZE]

    conn = get_db_connection()
    cursor = conn.cursor()
    placeholder = "%s" if DB_TYPE == "mysql" else "?"
    placeholders = ','.join(placeholder for _ in task_ids)
    exec_sql(cursor,
        f"SELECT id FROM collection_tasks WHERE id IN ({placeholders}) AND task_status = 0",
        task_ids)
    valid_ids = [row[0] if not isinstance(row, dict) else row['id'] for row in cursor.fetchall()]
    if valid_ids:
        up = ','.join(placeholder for _ in valid_ids)
        exec_sql(cursor,
            f"UPDATE collection_tasks SET task_status = 2, started_at = CURRENT_TIMESTAMP WHERE id IN ({up})",
            valid_ids)
    conn.commit()
    conn.close()

    if not valid_ids:
        return {"message": "没有待执行的任务", "started": 0}

    max_workers = 2
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        exec_sql(cursor, "SELECT enabled, max_concurrent FROM concurrency_settings WHERE id = 1")
        row = cursor.fetchone()
        conn.close()
        if row and (row[0] if not isinstance(row, dict) else row.get('enabled', 0)):
            max_workers = min((row[1] if not isinstance(row, dict) else row.get('max_concurrent', 2)), 3)
    except Exception:
        pass

    # ── 异步并发执行 ──
    semaphore = asyncio.Semaphore(max_workers)
    completed = 0
    failed = 0
    total = len(valid_ids)
    start_time = time.time()
    lock = threading.Lock()

    async def run_one(tid):
        nonlocal completed, failed
        async with semaphore:
            result = await process_single_task_async(tid)
        with lock:
            if result.get("status") == "success":
                completed += 1
                if result.get("extracted", 0) > 0:
                    print(f"[OK] 任务 {tid}: +{result['extracted']} 条")
            else:
                failed += 1
            done = completed + failed
            if done % 20 == 0:
                elapsed = time.time() - start_time
                rate = done / (elapsed / 60) if elapsed > 0 else 0
                print(f"[PROGRESS] {done}/{total} ({done*100//total}%) | 成功:{completed} 失败:{failed} | {rate:.1f}任务/分")

    async def run_all():
        await asyncio.gather(*[run_one(tid) for tid in valid_ids])

    threading.Thread(target=lambda: asyncio.run(run_all()), daemon=True).start()

    elapsed = time.time() - start_time
    return {
        "message": f"已启动 {len(valid_ids)} 个任务，并发数: {max_workers}",
        "started": len(valid_ids),
        "concurrent": max_workers,
        "task_ids": valid_ids[:10],
    }


@app.get("/api/settings/concurrency")
async def get_concurrency_settings():
    conn = get_db_connection()
    cursor = conn.cursor()
    exec_sql(cursor, "SELECT enabled, max_concurrent FROM concurrency_settings WHERE id = 1")
    row = cursor.fetchone()
    conn.close()
    if row:
        en = row[0] if not isinstance(row, dict) else row.get('enabled', 0)
        mc = row[1] if not isinstance(row, dict) else row.get('max_concurrent', 1)
        return {"success": True, "enabled": bool(en), "max_concurrent": int(mc)}
    return {"success": True, "enabled": False, "max_concurrent": 1}

@app.post("/api/settings/concurrency")
async def save_concurrency_settings(request: Request):
    data = await request.json()
    max_concurrent = min(max(int(data.get("max_concurrent", 1)), 1), 15)
    enabled = bool(data.get("enabled", False))
    conn = get_db_connection()
    cursor = conn.cursor()
    exec_sql(cursor, "UPDATE concurrency_settings SET enabled = ?, max_concurrent = ?, updated_at = CURRENT_TIMESTAMP WHERE id = 1", [int(enabled), max_concurrent])
    conn.commit()
    conn.close()
    write_log(username="system", operation_type="更新并发设置", operation_desc=f"enabled={enabled}, max_concurrent={max_concurrent}")
    return {"success": True, "enabled": enabled, "max_concurrent": max_concurrent}


@app.post("/api/collection-tasks/{task_id}/cancel")
async def cancel_task(task_id: int):
    """终止正在执行的任务 - 支持强制终止"""
    import threading
    
    conn = get_db_connection()
    cursor = conn.cursor()
    exec_sql(cursor, "SELECT task_status, retry_count FROM collection_tasks WHERE id = ?", [task_id])
    row = cursor.fetchone()
    if not row:
        conn.close()
        return {"success": False, "message": "任务不存在"}
    
    task_status = row[0]
    retry_count = row[1] or 0
    
    # 只能终止处理中(2)、等待中(0)或失败(4)的任务
    if task_status not in (2, 0, 4):
        conn.close()
        return {"success": False, "message": "只能终止处理中、等待中或失败的任务"}
    
    # 如果是等待中的任务且没有重试过，直接标记为已终止
    if task_status == 0 and retry_count <= 0:
        exec_sql(cursor, "UPDATE collection_tasks SET task_status = 5, completed_at = CURRENT_TIMESTAMP WHERE id = ?", [task_id])
        conn.commit()
        conn.close()
        return {"success": True, "message": "任务已终止"}
    
    # 设置取消信号，通知后台线程停止
    if hasattr(app, '_cancel_signals') and task_id in app._cancel_signals:
        app._cancel_signals[task_id].set()
    
    # 强制终止后台线程（如果存在）
    thread_key = f"task_thread_{task_id}"
    if hasattr(app, '_task_threads') and thread_key in app._task_threads:
        try:
            thread = app._task_threads[thread_key]
            if thread and thread.is_alive():
                # 注意：Python 没有安全的线程强制终止机制
                # 这里我们只是记录线程信息，主要依靠取消信号
                pass
        except:
            pass
        finally:
            # 清理线程记录
            del app._task_threads[thread_key]
    
    # 立即更新数据库状态为已终止
    exec_sql(cursor, "UPDATE collection_tasks SET task_status = 5, completed_at = CURRENT_TIMESTAMP, error_message = '任务被用户终止' WHERE id = ?", [task_id])
    conn.commit()
    conn.close()
    
    return {"success": True, "message": "任务已终止"}

@app.post("/api/collection-tasks/{task_id}/resume")
async def resume_task(task_id: int):
    """恢复已终止的任务"""
    conn = get_db_connection()
    cursor = conn.cursor()
    exec_sql(cursor, "SELECT task_status FROM collection_tasks WHERE id = ?", [task_id])
    row = cursor.fetchone()
    if not row:
        conn.close()
        return {"success": False, "message": "任务不存在"}
    if row[0] != 5:
        conn.close()
        return {"success": False, "message": "只能恢复已终止的任务"}
    exec_sql(cursor, "UPDATE collection_tasks SET task_status = 0, started_at = NULL, error_message = NULL WHERE id = ?", [task_id])
    conn.commit()
    conn.close()
    return {"success": True, "message": "任务已恢复，等待执行"}

@app.post("/api/collection-tasks/batch-delete")
async def batch_delete_tasks(request: Request):
    """批量删除采集任务"""
    data = await request.json()
    task_ids = data.get("task_ids", [])
    if not task_ids:
        raise HTTPException(status_code=400, detail="请提供要删除的任务 ID 列表")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        exec_sql(cursor, "PRAGMA foreign_keys = OFF")
        placeholders = ','.join('?' * len(task_ids))
        cursor.execute(f"DELETE FROM collection_tasks WHERE id IN ({placeholders})", task_ids)
        deleted_count = cursor.rowcount
        exec_sql(cursor, "PRAGMA foreign_keys = ON")
        conn.commit()
        conn.close()
        return {"success": True, "deleted": deleted_count, "message": f"已删除 {deleted_count} 个任务"}
    except Exception as e:
        conn.rollback()
        conn.close()
        raise HTTPException(status_code=500, detail=f"批量删除失败：{str(e)}")

@app.get("/api/settings/scheduler")
async def get_scheduler_settings():
    conn = get_db_connection()
    cursor = conn.cursor()
    exec_sql(cursor, "SELECT enabled, start_hour, end_hour, last_run_at FROM scheduler_settings WHERE id = 1")
    row = cursor.fetchone()
    conn.close()
    if row:
        if hasattr(row, 'get'):
            # 字典格式
            return {"success": True, "enabled": bool(row.get('enabled', 0)), "start_hour": row.get('start_hour', 1), "end_hour": row.get('end_hour', 5), "last_run_at": row.get('last_run_at')}
        else:
            # 元组格式
            return {"success": True, "enabled": bool(row[0]), "start_hour": row[1], "end_hour": row[2], "last_run_at": row[3]}
    return {"success": True, "enabled": False, "start_hour": 1, "end_hour": 5, "last_run_at": None}

@app.post("/api/settings/scheduler")
async def save_scheduler_settings(request: Request):
    data = await request.json()
    enabled = bool(data.get("enabled", False))
    start_hour = min(max(int(data.get("start_hour", 1)), 0), 23)
    end_hour = min(max(int(data.get("end_hour", 5)), 1), 24)
    conn = get_db_connection()
    cursor = conn.cursor()
    exec_sql(cursor, "UPDATE scheduler_settings SET enabled = ?, start_hour = ?, end_hour = ?, updated_at = CURRENT_TIMESTAMP WHERE id = 1", [int(enabled), start_hour, end_hour])
    conn.commit()
    conn.close()
    return {"success": True, "enabled": enabled, "start_hour": start_hour, "end_hour": end_hour}


@app.post("/api/collection-tasks/re-extract")
async def re_extract_all(request: Request):
    """重新提取：删除所有暂存数据 + 重置所有已完成任务为待处理"""
    conn = get_db_connection()
    cursor = conn.cursor()
    exec_sql(cursor, "DELETE FROM admission_records_staging")
    deleted_staging = cursor.rowcount
    exec_sql(cursor, "UPDATE collection_tasks SET task_status = 0, started_at = NULL, completed_at = NULL, error_message = NULL, retry_count = 0 WHERE task_status IN (2, 3, 4)")
    reset_count = cursor.rowcount
    conn.commit()
    conn.close()
    return {
        "success": True,
        "message": f"已清空 {deleted_staging} 条暂存记录，重置 {reset_count} 个任务为待处理",
        "deleted_staging": deleted_staging,
        "reset_tasks": reset_count,
    }


@app.get("/api/llm/config")
async def get_llm_config():
    """获取 LLM 审核配置"""
    from app.llm_screener import _load_config
    config = _load_config()
    config.pop("api_key", None)
    has_key = bool(_load_config().get("api_key"))
    return {"success": True, "config": config, "has_api_key": has_key}


@app.post("/api/llm/config")
async def save_llm_config(request: Request):
    """保存 LLM 审核配置"""
    from app.llm_screener import save_config, _load_config
    data = await request.json()
    cfg = {}
    for k in ("enabled", "provider", "api_base", "api_key", "model", "max_tokens", "temperature", "batch_size", "timeout", "corrector_enabled"):
        if k in data and data[k] is not None:
            cfg[k] = data[k]
    save_config(cfg)
    config = _load_config()
    config.pop("api_key", None)
    return {"success": True, "config": config, "has_api_key": bool(_load_config().get("api_key"))}


@app.get("/api/llm/stats")
async def get_llm_stats():
    """获取 LLM 审核统计"""
    from app.llm_screener import _ensure_tables
    _ensure_tables()
    conn = get_db_connection()
    cursor = conn.cursor()
    exec_sql(cursor, "SELECT COUNT(*) FROM ai_screen_stats")
    total_screens = cursor.fetchone()[0] or 0
    exec_sql(cursor, "SELECT COALESCE(SUM(total_candidates),0), COALESCE(SUM(passed),0), COALESCE(SUM(rejected),0) FROM ai_screen_stats")
    row = cursor.fetchone()
    total_in, total_pass, total_reject = row[0] or 0, row[1] or 0, row[2] or 0
    exec_sql(cursor, "SELECT COUNT(*) FROM ai_rejected_log")
    rejected_count = cursor.fetchone()[0] or 0
    conn.close()
    return {
        "success": True,
        "total_screen_events": total_screens,
        "total_candidates": total_in,
        "total_passed": total_pass,
        "total_rejected": total_reject,
        "rejected_log_count": rejected_count,
    }


@app.get("/api/admission-records/detail/{record_id}")
async def get_admission_record(record_id: int):
    """获取单条录取记录详情（优先主表，其次暂存表）"""
    conn = get_db_connection()
    cursor = conn.cursor()
    # 先查主表
    exec_sql(cursor, "SELECT * FROM admission_records WHERE id = ?", [record_id])
    row = cursor.fetchone()
    if not row:
        # 再查暂存表
        exec_sql(cursor, "SELECT * FROM admission_records_staging WHERE id = ?", [record_id])
        row = cursor.fetchone()
    conn.close()
    if not row:
        return {"success": False, "message": "记录不存在"}
    return {"success": True, "record": dict(row)}


@app.get("/api/stats/overview")
async def get_stats_overview():
    """总体统计：总数、国家数、大学数、奖学金总额"""
    from app.db_config import exec_sql
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 总记录数
    exec_sql(cursor, "SELECT COUNT(*) FROM admission_records")
    total_records = get_fetch_value(cursor.fetchone())
    
    # 涉及国家数（使用country文本字段）
    exec_sql(cursor, "SELECT COUNT(DISTINCT country) FROM admission_records WHERE country IS NOT NULL AND country != ''")
    total_countries = get_fetch_value(cursor.fetchone())
    
    # 录取大学数（使用university_cn字段）
    exec_sql(cursor, "SELECT COUNT(DISTINCT university_cn) FROM admission_records WHERE university_cn IS NOT NULL AND university_cn != ''")
    total_universities = get_fetch_value(cursor.fetchone())
    
    # 奖学金总额
    exec_sql(cursor, "SELECT COALESCE(SUM(scholarship_amount), 0) FROM admission_records")
    total_scholarship = get_fetch_value(cursor.fetchone())
    
    # 已验证记录数
    exec_sql(cursor, "SELECT COUNT(*) FROM admission_records WHERE status = 1")
    verified_count = get_fetch_value(cursor.fetchone())
    
    conn.close()
    
    return {
        "total_records": total_records,
        "total_countries": total_countries,
        "total_universities": total_universities,
        "total_scholarship": total_scholarship,
        "verified_count": verified_count
    }

@app.get("/api/stats/country")
async def get_stats_country():
    """按国家统计录取人数"""
    conn = get_db_connection()
    cursor = conn.cursor(pymysql.cursors.DictCursor) if DB_TYPE == "mysql" else conn.cursor()
    
    cursor.execute("SELECT country, COUNT(*) as count FROM admission_records WHERE country IS NOT NULL AND country != '' GROUP BY country ORDER BY count DESC")
    
    rows = cursor.fetchall()
    conn.close()
    
    return {"data": [{"country": row['country'], "count": row['count']} for row in rows]}

@app.get("/api/stats/university")
async def get_stats_university(top: int = 10):
    """按大学统计录取人数（TOP10）"""
    conn = get_db_connection()
    cursor = conn.cursor(pymysql.cursors.DictCursor) if DB_TYPE == "mysql" else conn.cursor()
    
    cursor.execute("SELECT university_cn, university_en, COUNT(*) as count FROM admission_records WHERE university_cn IS NOT NULL AND university_cn != '' GROUP BY university_cn, university_en ORDER BY count DESC LIMIT {}".format(top))
    
    rows = cursor.fetchall()
    conn.close()
    
    return {"data": [{"university": row['university_cn'], "university_en": row['university_en'], "count": row['count']} for row in rows]}

@app.get("/api/stats/trend")
async def get_stats_trend():
    """年度趋势统计"""
    conn = get_db_connection()
    cursor = conn.cursor(pymysql.cursors.DictCursor) if DB_TYPE == "mysql" else conn.cursor()
    
    cursor.execute("SELECT admission_year, COUNT(*) as count FROM admission_records WHERE admission_year IS NOT NULL GROUP BY admission_year ORDER BY admission_year ASC")
    
    rows = cursor.fetchall()
    conn.close()
    
    return {"data": [{"year": row['admission_year'], "count": row['count']} for row in rows]}

@app.get("/api/stats/major")
async def get_stats_major(top: int = 20):
    """按专业统计录取人数（TOP20）"""
    conn = get_db_connection()
    cursor = conn.cursor(pymysql.cursors.DictCursor) if DB_TYPE == "mysql" else conn.cursor()
    
    cursor.execute("SELECT major_cn, COUNT(*) as count FROM admission_records WHERE major_cn IS NOT NULL AND major_cn != '' GROUP BY major_cn ORDER BY count DESC LIMIT {}".format(top))
    
    rows = cursor.fetchall()
    conn.close()
    
    return {"data": [{"major": row['major_cn'], "count": row['count']} for row in rows]}

@app.get("/api/stats/school")
async def get_stats_school(top: int = 10):
    """按来源学校统计录取人数（TOP10）"""
    conn = get_db_connection()
    cursor = conn.cursor(pymysql.cursors.DictCursor) if DB_TYPE == "mysql" else conn.cursor()
    
    cursor.execute("SELECT source_school, COUNT(*) as count FROM admission_records WHERE source_school IS NOT NULL AND source_school != '' GROUP BY source_school ORDER BY count DESC, source_school ASC LIMIT {}".format(top))
    
    rows = cursor.fetchall()
    conn.close()
    
    return {"data": [
        {"rank": i + 1, "school": row['source_school'], "count": row['count']} 
        for i, row in enumerate(rows)
    ]}

# ============== 手动录入 API ==============

@app.get("/api/countries")
async def get_countries():
    """获取国家列表"""
    conn = get_db_connection()
    cursor = conn.cursor()
    exec_sql(cursor, "SELECT * FROM countries ORDER BY name_cn")
    rows = cursor.fetchall()
    conn.close()
    data = []
    for row in rows:
        d = dict(row)
        d['name'] = d.get('name_cn', '')
        data.append(d)
    return {"data": data}

@app.get("/api/universities")
async def get_universities(limit: int = 500):
    """获取大学列表"""
    conn = get_db_connection()
    cursor = conn.cursor()
    exec_sql(cursor, "SELECT * FROM universities ORDER BY name_cn LIMIT ?", [limit])
    rows = cursor.fetchall()
    conn.close()
    data = []
    for row in rows:
        d = dict(row)
        d['name'] = d.get('name_cn', '')
        data.append(d)
    return {"data": data}

@app.get("/api/source-schools")
async def get_source_schools():
    """获取来源学校列表"""
    conn = get_db_connection()
    cursor = conn.cursor()
    exec_sql(cursor, "SELECT * FROM source_schools ORDER BY school_name")
    rows = cursor.fetchall()
    conn.close()
    data = []
    for row in rows:
        d = dict(row)
        d['name'] = d.get('school_name', '')
        data.append(d)
    return {"data": data}

@app.post("/api/countries")
async def create_country(country: Dict[str, Any]):
    """创建新国家"""
    if "name" not in country:
        raise HTTPException(status_code=400, detail="缺少国家名称")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        exec_sql(cursor, "INSERT INTO countries (name, name_en) VALUES (?, ?)", [country.get("name"), country.get("name_en")])
        conn.commit()
        country_id = cursor.lastrowid
        exec_sql(cursor, "SELECT * FROM countries WHERE id = ?", [country_id])
        new_country = dict(cursor.fetchone())
        conn.close()
        return new_country
    except sqlite3.IntegrityError:
        conn.close()
        raise HTTPException(status_code=400, detail="国家已存在")

@app.post("/api/universities")
async def create_university(university: Dict[str, Any]):
    """创建新大学"""
    if "name" not in university:
        raise HTTPException(status_code=400, detail="缺少大学名称")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        exec_sql(cursor, "INSERT INTO universities (name, name_en, country_id) VALUES (?, ?, ?)", [university.get("name"), university.get("name_en"), university.get("country_id")])
        conn.commit()
        university_id = cursor.lastrowid
        exec_sql(cursor, "SELECT * FROM universities WHERE id = ?", [university_id])
        new_university = dict(cursor.fetchone())
        conn.close()
        return new_university
    except sqlite3.IntegrityError:
        conn.close()
        raise HTTPException(status_code=400, detail="大学已存在")

@app.post("/api/source-schools")
async def create_source_school(school: Dict[str, Any]):
    """创建新来源学校"""
    if "name" not in school:
        raise HTTPException(status_code=400, detail="缺少学校名称")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        exec_sql(cursor, "INSERT INTO source_schools (name, official_account) VALUES (?, ?)", [school.get("name"), school.get("official_account")])
        conn.commit()
        school_id = cursor.lastrowid
        exec_sql(cursor, "SELECT * FROM source_schools WHERE id = ?", [school_id])
        new_school = dict(cursor.fetchone())
        conn.close()
        return new_school
    except sqlite3.IntegrityError:
        conn.close()
        raise HTTPException(status_code=400, detail="学校已存在")

@app.post("/api/records/manual-entry")
async def manual_entry_record(record: Dict[str, Any]):
    """
    手动录入单条录取记录
    
    必填字段：student_name, university, country
    """
    # 验证必填字段
    required = ["student_name", "university", "country"]
    for field in required:
        if field not in record or not record[field]:
            raise HTTPException(status_code=400, detail=f"缺少必填字段：{field}")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # 1. 查找或创建国家
        country_name = record["country"]
        exec_sql(cursor, "SELECT id FROM countries WHERE name_cn = ?", [country_name])
        row = cursor.fetchone()
        if not row:
            exec_sql(cursor, "INSERT INTO countries (name_cn) VALUES (?)", [country_name])
        
        # 2. 查找或创建大学
        university_name = record["university"]
        exec_sql(cursor, "SELECT id FROM universities WHERE name_cn = ?", [university_name])
        row = cursor.fetchone()
        if not row:
            exec_sql(cursor, "INSERT INTO universities (name_cn) VALUES (?)", [university_name])
        
        # 3. 查找来源学校（可选）
        if record.get("source_school"):
            exec_sql(cursor, "SELECT id FROM source_schools WHERE school_name = ?", [record["source_school"]])
            row = cursor.fetchone()
            if not row:
                exec_sql(cursor, "INSERT INTO source_schools (school_name) VALUES (?)", [record["source_school"]])
        
        # 4. 插入录取记录（使用正确的字段名）
        cursor.execute("""
            INSERT INTO admission_records (
                student_name_cn, data_source, country, university_cn, major_cn,
                admission_type, admission_status, admission_year, scholarship_amount,
                scholarship_currency, notes, article_url, article_title,
                created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, [
            record.get("student_name"),
            record.get("data_source", ""),
            record.get("country"),
            record.get("university"),
            record.get("major_cn", record.get("major", "")),
            record.get("admission_type", ""),
            record.get("admission_status", "已录取"),
            record.get("admission_year"),
            record.get("scholarship_amount"),
            record.get("scholarship_currency", ""),
            record.get("notes", ""),
            record.get("article_url", ""),
            record.get("article_title", ""),
            get_local_time(),
            get_local_time()
        ])
        
        conn.commit()
        record_id = cursor.lastrowid
        
        # 返回创建的记录
        exec_sql(cursor, "SELECT * FROM admission_records WHERE id = ?", [record_id])
        new_record = dict(cursor.fetchone())
        conn.close()
        
        return {"success": True, "message": "录入成功", "record_id": record_id, "record": new_record}
        
    except sqlite3.IntegrityError as e:
        conn.close()
        raise HTTPException(status_code=400, detail=f"数据冲突：{str(e)}")
    except Exception as e:
        conn.close()

@app.post("/api/records/manual-entry/batch")
async def manual_entry_batch(batch_data: Dict[str, Any]):
    """
    批量手动录入录取记录
    
    请求格式：
    {
        "records": [
            {
                "student_name": "张三",
                "university": "UC Berkeley",
                "country": "美国",
                "major": "CS",
                "gpa": 3.8
            },
            ...
        ]
    }
    
    返回：
    {
        "success": true,
        "created": 10,
        "skipped": 2,
        "errors": []
    }
    """
    records = batch_data.get("records", [])
    if not records:
        raise HTTPException(status_code=400, detail="请提供记录数据")
    
    if len(records) > 100:
        raise HTTPException(status_code=400, detail="单次最多支持 100 条记录")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    created_count = 0
    skipped_count = 0
    errors = []
    
    for idx, record in enumerate(records):
        try:
            # 验证必填字段
            if not record.get("student_name") or not record.get("university") or not record.get("country"):
                errors.append(f"记录 {idx + 1}: 缺少必填字段")
                skipped_count += 1
                continue
            
            # 查找或创建国家（用于下拉列表）
            country_name = record["country"]
            exec_sql(cursor, "SELECT id FROM countries WHERE name_cn = ?", [country_name])
            row = cursor.fetchone()
            if not row:
                exec_sql(cursor, "INSERT INTO countries (name_cn) VALUES (?)", [country_name])
            
            # 查找或创建大学（用于下拉列表）
            university_name = record["university"]
            exec_sql(cursor, "SELECT id FROM universities WHERE name_cn = ?", [university_name])
            row = cursor.fetchone()
            if not row:
                exec_sql(cursor, "INSERT INTO universities (name_cn) VALUES (?)", [university_name])
            
            # 查找来源学校（可选）
            if record.get("source_school"):
                exec_sql(cursor, "SELECT id FROM source_schools WHERE school_name = ?", [record["source_school"]])
                row = cursor.fetchone()
                if not row:
                    exec_sql(cursor, "INSERT INTO source_schools (name) VALUES (?)", [record["source_school"]])
            
            # 插入录取记录（使用现有表结构的文本字段）
            cursor.execute("""
                INSERT INTO admission_records (
                    student_name, student_name_en, source_school, country, university, university_en,
                    major, major_en, gpa, toefl, ielts, sat,
                    scholarship, scholarship_amount, scholarship_type, degree,
                    admission_year, article_url, article_title, requirements, portfolio,
                    is_verified, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            """, [
                record.get("student_name"),
                record.get("student_name_en"),
                record.get("source_school"),
                record.get("country"),
                record.get("university"),
                record.get("university_en"),
                record.get("major"),
                record.get("major_en"),
                record.get("gpa"),
                record.get("toefl"),
                record.get("ielts"),
                record.get("sat"),
                record.get("scholarship"),
                record.get("scholarship_amount"),
                record.get("scholarship"),  # scholarship_type
                record.get("degree"),
                record.get("admission_year"),
                record.get("article_url"),
                record.get("article_title"),
                record.get("requirements"),
                record.get("portfolio"),
                1 if record.get("is_verified") else 0
            ])
            
            created_count += 1
            
        except sqlite3.IntegrityError as e:
            errors.append(f"记录 {idx + 1}: 数据冲突 - {str(e)}")
            skipped_count += 1
        except Exception as e:
            errors.append(f"记录 {idx + 1}: {str(e)}")
            skipped_count += 1
    
    conn.commit()
    conn.close()
    
    return {
        "success": len(errors) == 0,
        "created": created_count,
        "skipped": skipped_count,
        "errors": errors[:10]  # 最多返回 10 个错误
    }

@app.get("/api/filters/options")
async def get_filter_options():
    """
    获取筛选器所需的所有选项（来源学校、国家、年份、验证状态）
    """
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 获取来源学校列表
    exec_sql(cursor, "SELECT DISTINCT ss.school_name FROM source_schools ss INNER JOIN admission_records ar ON ss.id = ar.source_school_id WHERE ss.school_name IS NOT NULL ORDER BY ss.name")
    source_schools = [row[0] for row in cursor.fetchall()]
    
    # 获取国家列表
    exec_sql(cursor, "SELECT DISTINCT c.name_cn FROM countries c INNER JOIN admission_records ar ON c.id = ar.country_id WHERE c.name_cn IS NOT NULL ORDER BY c.name")
    countries = [row[0] for row in cursor.fetchall()]
    
    # 获取年份列表（从大到小）
    exec_sql(cursor, "SELECT DISTINCT admission_year FROM admission_records WHERE admission_year IS NOT NULL ORDER BY admission_year DESC")
    years = [row[0] for row in cursor.fetchall()]
    
    # 验证状态选项
    verified_options = [
        {"value": "", "label": "全部"},
        {"value": "true", "label": "已验证"},
        {"value": "false", "label": "未验证"}
    ]
    
    conn.close()
    
    return {
        "source_schools": source_schools,
        "countries": countries,
        "years": years,
        "verified": verified_options
    }

# ============== Excel 导入 API ==============

@app.get("/api/import/template")
async def download_import_template():
    """
    下载 Excel 导入模板
    """
    template_path = os.path.join(os.path.dirname(__file__), "templates", "录取数据导入模板.xlsx")
    
    if not os.path.exists(template_path):
        raise HTTPException(status_code=404, detail="模板文件不存在")
    
    import urllib.parse
    filename_encoded = urllib.parse.quote("录取数据导入模板.xlsx")
    
    return FileResponse(
        template_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}"
        }
    )


@app.post("/api/import/excel")
async def import_excel(file: UploadFile = File(...)):
    """
    从 Excel 文件批量导入录取数据
    
    支持字段：
    - 必填：student_name(学生姓名), country(国家), university(录取大学)
    - 可选：student_name_en, source_school, university_en, major, major_en, degree,
           gpa, toefl, ielts, sat, act, gre, gmat, scholarship, scholarship_amount,
           application_year, admission_year, article_url, article_title,
           requirements, portfolio, is_verified
    
    返回：
    {
        "success": true/false,
        "total": 总记录数，
        "created": 成功导入数，
        "skipped": 跳过数（重复数据）,
        "errors": [错误详情列表]
    }
    """
    # 验证文件类型
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="请上传 Excel 文件 (.xlsx 或 .xls)")
    
    # 限制文件大小 (10MB)
    file_size = 0
    file_content = await file.read()
    file_size = len(file_content)
    if file_size > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="文件大小不能超过 10MB")
    
    try:
        # 读取 Excel 文件
        df = pd.read_excel(io.BytesIO(file_content), sheet_name="填写示例")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法读取 Excel 文件：{str(e)}")
    
    if df.empty:
        raise HTTPException(status_code=400, detail="Excel 文件为空")
    
    # 限制导入数量
    if len(df) > 1000:
        raise HTTPException(status_code=400, detail="单次最多支持导入 1000 条记录")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    created_count = 0
    skipped_count = 0
    errors = []
    
    # 预加载现有数据用于重复检测
    cursor.execute("""
        SELECT student_name, university, admission_year 
        FROM admission_records 
        WHERE student_name IS NOT NULL AND university IS NOT NULL
    """)
    existing_records = set()
    for row in cursor.fetchall():
        existing_records.add((row[0], row[1], str(row[2]) if row[2] else ""))
    
    for idx, row in df.iterrows():
        try:
            row_num = idx + 2  # Excel 行号（从 2 开始，因为第 1 行是表头）
            
            # 1. 验证必填字段
            student_name = str(row.get('student_name', '')).strip() if pd.notna(row.get('student_name')) else ''
            country = str(row.get('country', '')).strip() if pd.notna(row.get('country')) else ''
            university = str(row.get('university', '')).strip() if pd.notna(row.get('university')) else ''
            
            if not student_name:
                errors.append(f"第{row_num}行：缺少必填字段 - 学生姓名 (student_name)")
                skipped_count += 1
                continue
            
            if not country:
                errors.append(f"第{row_num}行：缺少必填字段 - 国家/地区 (country)")
                skipped_count += 1
                continue
            
            if not university:
                errors.append(f"第{row_num}行：缺少必填字段 - 录取大学 (university)")
                skipped_count += 1
                continue
            
            # 2. 重复检测（学生姓名 + 大学 + 入学年份）
            admission_year = str(int(row.get('admission_year', 2026))) if pd.notna(row.get('admission_year')) else "2026"
            record_key = (student_name, university, admission_year)
            
            if record_key in existing_records:
                errors.append(f"第{row_num}行：重复数据 - {student_name} + {university} + {admission_year}")
                skipped_count += 1
                continue
            
            # 3. 准备其他字段
            def safe_float(val, default=None):
                try:
                    return float(val) if pd.notna(val) else default
                except (ValueError, TypeError):
                    return default
            
            def safe_int(val, default=None):
                try:
                    return int(float(val)) if pd.notna(val) else default
                except (ValueError, TypeError):
                    return default
            
            # 4. 插入记录（使用文本字段而非外键 ID）
            cursor.execute("""
                INSERT INTO admission_records (
                    student_name, student_name_en, source_school, country, university,
                    university_en, major, major_en, degree, gpa, toefl, ielts, sat, act, gre, gmat,
                    scholarship, scholarship_amount, scholarship_type, application_year, admission_year,
                    article_url, article_title, requirements, portfolio,
                    is_verified, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            """, [
                student_name,
                str(row.get('student_name_en', '')).strip() if pd.notna(row.get('student_name_en')) else None,
                str(row.get('source_school', '')).strip() if pd.notna(row.get('source_school')) else None,
                country,
                university,
                str(row.get('university_en', '')).strip() if pd.notna(row.get('university_en')) else None,
                str(row.get('major', '')).strip() if pd.notna(row.get('major')) else None,
                str(row.get('major_en', '')).strip() if pd.notna(row.get('major_en')) else None,
                str(row.get('degree', '')).strip() if pd.notna(row.get('degree')) else None,
                safe_float(row.get('gpa')),
                safe_int(row.get('toefl')),
                safe_float(row.get('ielts')),
                safe_int(row.get('sat')),
                safe_int(row.get('act')),
                safe_int(row.get('gre')),
                safe_int(row.get('gmat')),
                str(row.get('scholarship', '')).strip() if pd.notna(row.get('scholarship')) else None,
                safe_float(row.get('scholarship_amount')),
                str(row.get('scholarship', '')).strip() if pd.notna(row.get('scholarship')) else None,  # scholarship_type
                safe_int(row.get('application_year')),
                safe_int(row.get('admission_year')),
                str(row.get('article_url', '')).strip() if pd.notna(row.get('article_url')) else None,
                str(row.get('article_title', '')).strip() if pd.notna(row.get('article_title')) else None,
                str(row.get('requirements', '')).strip() if pd.notna(row.get('requirements')) else None,
                str(row.get('portfolio', '')).strip() if pd.notna(row.get('portfolio')) else None,
                1 if safe_int(row.get('is_verified'), 0) else 0
            ])
            
            created_count += 1
            existing_records.add(record_key)  # 添加到已存在集合，防止同一文件中重复
            
        except Exception as e:
            errors.append(f"第{idx + 2}行：{str(e)}")
            skipped_count += 1
    
    conn.commit()
    conn.close()
    
    return {
        "success": len(errors) == 0,
        "total": len(df),
        "created": created_count,
        "skipped": skipped_count,
        "errors": errors[:20]  # 最多返回 20 个错误
    }# ==================== 导出功能（2026-05-08 修复） ====================

@app.post("/api/records/export/pdf")
async def export_records_pdf(filters: Dict[str, Any] = None):
    """导出录取记录为 PDF 格式报告"""
    if filters is None:
        filters = {}
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 构建筛选条件
    where_clauses = []
    params = []
    
    if filters.get('country'):
        where_clauses.append(f"country = {placeholder}")
        params.append(filters['country'])
    if filters.get('university'):
        where_clauses.append("university_cn LIKE ?")
        params.append(f"%{filters['university']}%")
    if filters.get('major'):
        where_clauses.append("major_cn LIKE ?")
        params.append(f"%{filters['major']}%")
    if filters.get('source_school'):
        where_clauses.append("source_school LIKE ?")
        params.append(f"%{filters['source_school']}%")
    if filters.get('year_start'):
        where_clauses.append("admission_year >= ?")
        params.append(filters['year_start'])
    if filters.get('year_end'):
        where_clauses.append("admission_year <= ?")
        params.append(filters['year_end'])
    if filters.get('is_verified') is not None:
        where_clauses.append("is_verified = ?")
        params.append(1 if filters['is_verified'] else 0)
    
    where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""
    
    query = f"""
    SELECT * FROM admission_records
    {where_sql}
    ORDER BY admission_year DESC, created_at DESC
    """
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    
    records = [dict(row) for row in rows]
    
    if not records:
        raise HTTPException(status_code=404, detail="当前筛选条件下无数据可导出")
    
    try:
        from urllib.parse import quote
        from app.pdf_template import create_pdf_report
        pdf_data = create_pdf_report(records, filters)
        date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"录取数据_{date_str}.pdf"
        
        return Response(
            content=pdf_data,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF 生成失败：{str(e)}")


@app.post("/api/records/export/excel")
async def export_records_excel(filters: Dict[str, Any] = None):
    """导出录取记录为 Excel 格式"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="缺少 openpyxl 依赖")
    
    if filters is None:
        filters = {}
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 构建筛选条件
    where_clauses = []
    params = []
    
    if filters.get('country'):
        where_clauses.append(f"country = {placeholder}")
        params.append(filters['country'])
    if filters.get('university'):
        where_clauses.append("university_cn LIKE ?")
        params.append(f"%{filters['university']}%")
    if filters.get('major'):
        where_clauses.append("major_cn LIKE ?")
        params.append(f"%{filters['major']}%")
    if filters.get('source_school'):
        where_clauses.append("source_school LIKE ?")
        params.append(f"%{filters['source_school']}%")
    if filters.get('year_start'):
        where_clauses.append("admission_year >= ?")
        params.append(filters['year_start'])
    if filters.get('year_end'):
        where_clauses.append("admission_year <= ?")
        params.append(filters['year_end'])
    if filters.get('is_verified') is not None:
        where_clauses.append("is_verified = ?")
        params.append(1 if filters['is_verified'] else 0)
    if filters.get('student_name'):
        where_clauses.append("student_name_cn LIKE ?")
        params.append(f"%{filters['student_name']}%")
    
    where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""
    
    query = f"""
    SELECT * FROM admission_records
    {where_sql}
    ORDER BY admission_year DESC, created_at DESC
    """
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    
    if not rows:
        raise HTTPException(status_code=404, detail="当前筛选条件下无数据可导出")
    
    # 获取用户选择的字段
    selected_fields = filters.get('fields', [])
    if not selected_fields:
        # 默认导出必填字段
        selected_fields = ['student_name_cn', 'country', 'university_cn']
    
    # 字段映射 (key -> 中文名称)
    field_labels = {
        'student_name_cn': '学生姓名',
        'student_name_en': '英文名',
        'student_grade': '年级',
        'source_school': '来源学校',
        'country': '国家',
        'country_en': '国家（英文）',
        'university_cn': '录取大学',
        'university_en': '大学（英文）',
        'university_type': '大学类型',
        'university_ranking': '大学排名',
        'major_cn': '专业',
        'major_en': '专业（英文）',
        'major_category': '专业类别',
        'degree': '学位',
        'admission_type': '录取类型',
        'admission_status': '录取状态',
        'conditional_offer': '条件录取',
        'admission_date': '录取日期',
        'admission_year': '录取年份',
        'language_requirement_type': '语言考试类型',
        'language_score_required': '语言分数要求',
        'language_score_actual': '语言实际分数',
        'language_waived': '语言豁免',
        'sat_required': 'SAT要求',
        'sat_actual': 'SAT实际分数',
        'test_optional': 'Test Optional',
        'scholarship_amount': '奖学金金额',
        'scholarship_currency': '货币单位',
        'scholarship_type': '奖学金类型',
        'gpa': 'GPA',
        'toefl': 'TOEFL',
        'ielts': 'IELTS',
        'sat': 'SAT',
        'act': 'ACT',
        'gre': 'GRE',
        'gmat': 'GMAT',
        'article_url': '文章链接',
        'article_title': '文章标题',
        'publish_date': '发布日期',
        'is_verified': '验证状态',
        'created_at': '创建时间',
        'updated_at': '更新时间',
        'notes': '备注',
    }
    
    # 过滤存在的字段
    columns = [(field_labels.get(f, f), f) for f in selected_fields if f in field_labels]
    if not columns:
        columns = [(field_labels.get(f, f), f) for f in ['student_name_cn', 'country', 'university_cn']]
    
    # 创建 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "录取数据"
    
    header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1a73e8', end_color='1a73e8', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='center', wrap_text=True)
    
    # 表头
    for col_idx, (header, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 数据
    for row_idx, row in enumerate(rows, 2):
        d = dict(row)
        for col_idx, (_, field) in enumerate(columns, 1):
            value = d.get(field, '')
            if field == 'is_verified':
                value = '✅ 已验证' if value else '⏳ 待验证'
            elif field in ('created_at', 'updated_at', 'admission_date', 'publish_date') and value:
                value = str(value)[:19]
            elif field == 'conditional_offer':
                value = '是' if value == 1 else ('否' if value == 0 else '')
            elif field == 'language_waived':
                value = '是' if value == 1 else ('否' if value == 0 else '')
            elif field == 'test_optional':
                value = '是' if value == 1 else ('否' if value == 0 else '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
    
    # 列宽
    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18
    
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    now = datetime.now().strftime('%Y%m%d_%H%M%S')
    from urllib.parse import quote
    filename = f"录取数据_{now}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    )


@app.post("/api/records/export/csv")
async def export_records_csv(filters: Dict[str, Any] = None):
    """导出录取记录为 CSV 格式"""
    import csv
    import io
    
    if filters is None:
        filters = {}
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 构建筛选条件
    where_clauses = []
    params = []
    
    if filters.get('country'):
        where_clauses.append(f"country = {placeholder}")
        params.append(filters['country'])
    if filters.get('university'):
        where_clauses.append("university_cn LIKE ?")
        params.append(f"%{filters['university']}%")
    if filters.get('major'):
        where_clauses.append("major_cn LIKE ?")
        params.append(f"%{filters['major']}%")
    if filters.get('source_school'):
        where_clauses.append("source_school LIKE ?")
        params.append(f"%{filters['source_school']}%")
    if filters.get('year_start'):
        where_clauses.append("admission_year >= ?")
        params.append(filters['year_start'])
    if filters.get('year_end'):
        where_clauses.append("admission_year <= ?")
        params.append(filters['year_end'])
    if filters.get('is_verified') is not None:
        where_clauses.append("is_verified = ?")
        params.append(1 if filters['is_verified'] else 0)
    if filters.get('student_name'):
        where_clauses.append("student_name_cn LIKE ?")
        params.append(f"%{filters['student_name']}%")
    
    where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""
    
    query = f"""
    SELECT * FROM admission_records
    {where_sql}
    ORDER BY admission_year DESC, created_at DESC
    """
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    
    if not rows:
        raise HTTPException(status_code=404, detail="当前筛选条件下无数据可导出")
    
    # 获取用户选择的字段
    selected_fields = filters.get('fields', [])
    if not selected_fields:
        selected_fields = ['student_name_cn', 'country', 'university_cn']
    
    # 字段映射
    field_labels = {
        'student_name_cn': '学生姓名',
        'student_name_en': '英文名',
        'student_grade': '年级',
        'source_school': '来源学校',
        'country': '国家',
        'country_en': '国家（英文）',
        'university_cn': '录取大学',
        'university_en': '大学（英文）',
        'university_type': '大学类型',
        'university_ranking': '大学排名',
        'major_cn': '专业',
        'major_en': '专业（英文）',
        'major_category': '专业类别',
        'degree': '学位',
        'admission_type': '录取类型',
        'admission_status': '录取状态',
        'conditional_offer': '条件录取',
        'admission_date': '录取日期',
        'admission_year': '录取年份',
        'language_requirement_type': '语言考试类型',
        'language_score_required': '语言分数要求',
        'language_score_actual': '语言实际分数',
        'language_waived': '语言豁免',
        'sat_required': 'SAT要求',
        'sat_actual': 'SAT实际分数',
        'test_optional': 'Test Optional',
        'scholarship_amount': '奖学金金额',
        'scholarship_currency': '货币单位',
        'scholarship_type': '奖学金类型',
        'gpa': 'GPA',
        'toefl': 'TOEFL',
        'ielts': 'IELTS',
        'sat': 'SAT',
        'act': 'ACT',
        'gre': 'GRE',
        'gmat': 'GMAT',
        'article_url': '文章链接',
        'article_title': '文章标题',
        'publish_date': '发布日期',
        'is_verified': '验证状态',
        'created_at': '创建时间',
        'updated_at': '更新时间',
        'notes': '备注',
    }
    
    columns = [(field_labels.get(f, f), f) for f in selected_fields if f in field_labels]
    if not columns:
        columns = [(field_labels.get(f, f), f) for f in ['student_name_cn', 'country', 'university_cn']]
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([col[0] for col in columns])
    
    for row in rows:
        d = dict(row)
        row_data = []
        for _, field in columns:
            value = d.get(field, '')
            if field == 'is_verified':
                value = '已验证' if value else '待验证'
            elif field in ('created_at', 'updated_at', 'admission_date', 'publish_date') and value:
                value = str(value)[:19]
            elif field in ('conditional_offer', 'language_waived', 'test_optional'):
                value = '是' if value == 1 else ('否' if value == 0 else '')
            row_data.append(str(value) if value is not None else '')
        writer.writerow(row_data)
    
    csv_content = '\ufeff' + output.getvalue()
    now = datetime.now().strftime('%Y%m%d_%H%M%S')
    from urllib.parse import quote
    filename = f"录取数据_{now}.csv"
    
    return Response(
        content=csv_content.encode('utf-8-sig'),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    )


# ==================== 图片识别提取 API（方案 A：2026-05-09 新增） ====================

@app.post("/api/records/extract-from-image")
async def extract_from_image(request: Request):
    """
    通过 URL 打开文章 → 截图 → AI 视觉模型识别 → 提取录取字段
    
    请求格式：
    {
        "url": "https://mp.weixin.qq.com/s/xxx",
        "timeout": 30  (可选，默认 30 秒)
    }
    
    返回格式：
    {
        "success": true,
        "page_title": "2026届录取喜报",
        "page_url": "实际 URL",
        "records": [
            {
                "student_name_cn": "张三",
                "university_cn": "牛津大学",
                ...
            }
        ],
        "images_analyzed": 3,
        "error": ""
    }
    """
    data = await request.json()
    url = data.get("url", "").strip()
    timeout = data.get("timeout", 30)
    
    if not url:
        raise HTTPException(status_code=400, detail="请提供文章 URL")
    
    if not url.startswith(("http://", "https://")):
        url = "https://" + url
    
    try:
        from app.image_extractor import extract_from_url
        result = extract_from_url(url, timeout)
        return result
    except ImportError as e:
        raise HTTPException(status_code=500, detail=f"图片识别模块未安装: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"识别失败: {str(e)}")


@app.post("/api/records/save-extracted")
async def save_extracted_records(request: Request):
    """
    保存从图片识别提取的录取记录到数据库
    
    请求格式：
    {
        "records": [
            {
                "student_name_cn": "张三",
                "university_cn": "牛津大学",
                "country": "英国",
                ...
            }
        ]
    }
    """
    data = await request.json()
    records = data.get("records", [])
    
    if not records:
        raise HTTPException(status_code=400, detail="请提供要保存的记录")
    
    if len(records) > 100:
        raise HTTPException(status_code=400, detail="单次最多保存 100 条记录")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    saved = 0
    skipped = 0
    errors = []
    
    # 允许保存的字段
    allowed_fields = [
        "student_name_cn", "student_name_en", "student_grade",
        "country", "country_en",
        "university_cn", "university_en", "university_type", "university_ranking",
        "source_school",
        "major_cn", "major_en", "major_category",
        "admission_type", "admission_status", "conditional_offer",
        "admission_date", "admission_year",
        "language_requirement_type", "language_score_required",
        "language_score_actual", "language_waived",
        "sat_required", "sat_actual", "test_optional",
        "scholarship_amount", "scholarship_currency", "scholarship_type",
        "article_url", "article_title", "publish_date",
        "recognition_model",
        "notes"
    ]
    
    for idx, record in enumerate(records):
        try:
            # 过滤掉学生姓名为'未知'的记录
            if record.get("student_name_cn") == "未知":
                skipped += 1
                continue
            
            # 验证必填字段
            if not record.get("student_name_cn") or not record.get("university_cn") or not record.get("country"):
                errors.append(f"记录 {idx + 1}: 缺少必填字段（学生姓名/大学/国家）")
                skipped += 1
                continue
            
            # 过滤只保留允许的字段
            clean_record = {k: v for k, v in record.items() if k in allowed_fields and v and v != "未知"}
            
            # 处理 admission_year 为整数
            if "admission_year" in clean_record:
                try:
                    clean_record["admission_year"] = int(clean_record["admission_year"])
                except (ValueError, TypeError):
                    clean_record["admission_year"] = 2026
            
            # 处理 scholarship_amount 为浮点数
            if "scholarship_amount" in clean_record:
                try:
                    # 清理货币符号和逗号
                    amt = str(clean_record["scholarship_amount"]).replace(",", "").replace("$", "").replace("£", "").replace("€", "").strip()
                    clean_record["scholarship_amount"] = float(amt)
                except (ValueError, TypeError):
                    del clean_record["scholarship_amount"]
            
            # 构建插入 SQL
            columns = list(clean_record.keys())
            if "created_at" not in columns:
                columns.append("created_at")
                clean_record["created_at"] = "CURRENT_TIMESTAMP"
            columns.append("updated_at")
            
            placeholders = []
            values = []
            for col in columns:
                if col in ("created_at", "updated_at"):
                    placeholders.append("CURRENT_TIMESTAMP")
                else:
                    placeholders.append("?")
                    values.append(clean_record.get(col))
            
            query = f"INSERT INTO admission_records_staging ({', '.join(columns)}, review_status) VALUES ({', '.join(placeholders)}, 'pending')"
            cursor.execute(query, values)
            saved += 1
            
        except Exception as e:
            errors.append(f"记录 {idx + 1}: {str(e)}")
            skipped += 1
    
    conn.commit()
    conn.close()
    
    return {
        "success": len(errors) == 0 or saved > 0,
        "saved": saved,
        "skipped": skipped,
        "errors": errors[:10]
    }


# ==================== 免费公共图片识别模型 ====================

RECOGNITION_MODELS = [
    {
        "id": "qwen3.5-plus",
        "name": "通义千问 3.5 Plus",
        "description": "多模态模型，综合性能均衡，推荐使用",
        "provider": "DashScope (阿里云)",
        "api_base": "https://dashscope.aliyuncs.com/compatible-mode/v1",
        "free": True,
        "recommended": True
    },
    {
        "id": "qwen3.6-plus",
        "name": "通义千问 3.6 Plus",
        "description": "最新多模态模型，识别精度更高",
        "provider": "DashScope (阿里云)",
        "api_base": "https://dashscope.aliyuncs.com/compatible-mode/v1",
        "free": True,
        "recommended": False
    }
]


@app.get("/api/recognition-models")
async def get_recognition_models():
    """获取可用的免费图片识别模型列表"""
    return {"models": RECOGNITION_MODELS}


@app.post("/api/records/recognize-from-upload")
async def recognize_from_image_upload(
    model: str = Form("qwen3.5-plus"),
    file: UploadFile = File(...)
):
    """
    上传图片 → 使用指定模型识别 → 返回结构化录取记录

    支持 multipart/form-data:
    - file: 图片文件 (png/jpg/jpeg/webp)
    - model: 模型 ID (默认 qwen-vl-max)
    """
    import base64
    from PIL import Image

    if not file:
        raise HTTPException(status_code=400, detail="请上传图片文件")

    # 验证模型
    model_ids = [m["id"] for m in RECOGNITION_MODELS]
    if model not in model_ids:
        raise HTTPException(status_code=400, detail=f"不支持的模型: {model}，可选: {', '.join(model_ids)}")

    # 获取模型配置（"qwen3.6-plus" 使用旧的 API Base）
    model_config = next((m for m in RECOGNITION_MODELS if m["id"] == model), RECOGNITION_MODELS[0])

    # 读取上传文件
    contents = await file.read()
    if len(contents) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="图片大小不能超过 20MB")

    try:
        img = Image.open(io.BytesIO(contents))
        img.verify()
    except Exception:
        raise HTTPException(status_code=400, detail="无效的图片文件，请上传 PNG/JPG/JPEG/WEBP 格式")

    # 重新打开（verify 后需要重新加载）
    img = Image.open(io.BytesIO(contents))
    # 缩放大图片以加快识别
    max_dim = 2048
    if img.width > max_dim or img.height > max_dim:
        ratio = max_dim / max(img.width, img.height)
        img = img.resize((int(img.width * ratio), int(img.height * ratio)), Image.LANCZOS)

    buffered = io.BytesIO()
    img.save(buffered, format="PNG")
    b64 = base64.b64encode(buffered.getvalue()).decode()
    b64_url = f"data:image/png;base64,{b64}"

    # 选择合适的 API Key
    if model == "qwen3.6-plus":
        api_key = "sk-sp-61e3d323b258454fb3d6b67a50144d29"
    else:
        api_key = "sk-sp-61e3d323b258454fb3d6b67a50144d29"

    api_base = model_config["api_base"]

    try:
        client = OpenAI(api_key=api_key, base_url=api_base)

        prompt = """你是一个专业的留学录取信息提取助手。请仔细识别图片中的所有录取信息，包括：
- 学生姓名（中文和英文）
- 录取大学（中文和英文）
- 专业（中文和英文）
- 国家/地区
- 录取年份
- 奖学金金额和货币
- 录取类型/状态
- 来源学校（高中）

请以 JSON 数组格式返回，每条记录一个对象，字段名使用：
student_name_cn, student_name_en, university_cn, university_en,
major_cn, major_en, country, admission_year, scholarship_amount,
scholarship_currency, admission_type, admission_status, source_school

如果某些字段不确定，填"未知"。只返回 JSON 数组，不要其他文字。"""

        resp = client.chat.completions.create(
            model=model,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": b64_url}}
                ]
            }],
            max_tokens=4000,
            temperature=0.1
        )
        text = resp.choices[0].message.content

        json_match = re.search(r'\[[\s\S]*\]', text)
        if not json_match:
            return {
                "success": False,
                "records": [],
                "model_used": model,
                "raw_response": text[:500],
                "error": "未能从图片中提取到有效的录取信息，请确认图片包含录取相关内容"
            }

        records = json.loads(json_match.group())
        if isinstance(records, dict):
            records = [records]

        for r in records:
            r["recognition_model"] = model

        return {
            "success": len(records) > 0,
            "records": records,
            "model_used": model,
            "error": "" if records else "未能提取到有效记录"
        }

    except Exception as e:
        logger = logging.getLogger('ImageRecognizer')
        logger.error(f"图片识别失败(model={model}): {e}")
        raise HTTPException(status_code=500, detail=f"图片识别失败: {str(e)}")


# ==================== 数据审核 API（2026-05-09 新增） ====================

@app.get("/api/review/pending")
async def get_pending_reviews(status: str = None, page: int = Query(1, ge=1), page_size: int = Query(20, ge=1)):
    """获取待审核记录列表"""
    from app.db_config import DB_TYPE, fetch_count
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    placeholder = "%s" if DB_TYPE == "mysql" else "?"
    
    query = """SELECT id, student_name_cn, student_name_en, university_cn, major_cn, 
                      country, admission_year, scholarship_amount, scholarship_currency,
                      source_school, article_url, article_title,
                      review_status, review_comment, reviewed_by, reviewed_at,
                      recognition_model,
                      created_at, updated_at, data_source, promoted_at
               FROM admission_records_staging"""
    
    where_clauses = []
    params = []
    
    if status and status != 'all':
        where_clauses.append(f"review_status = {placeholder}")
        params.append(status)
    
    if where_clauses:
        query += " WHERE " + " AND ".join(where_clauses)
    
    query += " ORDER BY created_at DESC"
    
    # 获取总数
    count_query = """SELECT COUNT(*) FROM admission_records_staging"""
    if where_clauses:
        count_query += " WHERE " + " AND ".join(where_clauses)
    total = fetch_count(cursor, count_query, params if where_clauses else None)
    
    # 分页
    offset = (page - 1) * page_size
    query += " LIMIT {} OFFSET {}".format(page_size, offset)
    if params:
        cursor.execute(query, params)
    else:
        cursor.execute(query)
    rows = cursor.fetchall()
    records = [dict(row) for row in rows]
    
    # 格式化时间
    for r in records:
        if r.get('created_at'):
            r['created_at'] = str(r['created_at']).replace('T', ' ')[:19]
        if r.get('reviewed_at'):
            r['reviewed_at'] = str(r['reviewed_at']).replace('T', ' ')[:19]
    
    conn.close()
    
    # 确保 page_size 不为 0
    if page_size <= 0:
        page_size = 20
    return {
        "records": records,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size if page_size > 0 else 0
    }


@app.post("/api/review/batch")
async def batch_review(request: Dict[str, Any]):
    """批量审核（暂存表→审核通过→迁入主表）"""
    record_ids = request.get("record_ids", [])
    action = request.get("action")  # approved / rejected
    comment = request.get("comment", "")
    reviewer = request.get("reviewer", "系统")
    
    if not record_ids:
        raise HTTPException(status_code=400, detail="请选择要审核的记录")
    
    if action not in ("approved", "rejected"):
        raise HTTPException(status_code=400, detail="审核动作必须是 approved 或 rejected")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    ph = "%s" if DB_TYPE == "mysql" else "?"
    
    approved = 0
    rejected = 0
    
    for record_id in record_ids:
        exec_sql(cursor, "SELECT * FROM admission_records_staging WHERE id = " + ph, [record_id])
        row = cursor.fetchone()
        if not row:
            continue
        
        staging = dict(row)
        
        exec_sql(cursor,
            "UPDATE admission_records_staging SET review_status = " + ph + ", review_comment = " + ph + ", reviewed_by = " + ph + ", reviewed_at = CURRENT_TIMESTAMP WHERE id = " + ph,
            [action, comment, reviewer, record_id])
        
        if action == "approved":
            try:
                ph = "%s" if DB_TYPE == "mysql" else "?"
                cols = (
                    "student_name_cn, student_name_en, student_grade, "
                    "country, country_en, university_cn, university_en, "
                    "university_type, university_ranking, major_cn, major_en, "
                    "major_category, admission_type, admission_status, "
                    "admission_date, admission_year, language_requirement_type, "
                    "language_score_required, sat_required, "
                    "scholarship_amount, scholarship_currency, scholarship_type, "
                    "source_school, article_url, article_title, "
                    "recognition_model, data_source, "
                    "status, review_status, review_note, reviewed_by, reviewed_at, promoted_at, "
                    "created_at, updated_at, notes"
                )
                placeholders = (ph + ", ") * 35 + ph
                values = [
                    staging.get("student_name_cn") or "", staging.get("student_name_en") or "", staging.get("student_grade") or "",
                    staging.get("country") or "", staging.get("country_en") or "", staging.get("university_cn") or "", staging.get("university_en") or "",
                    staging.get("university_type") or "", staging.get("university_ranking") if staging.get("university_ranking") is not None else -1, staging.get("major_cn") or "", staging.get("major_en") or "",
                    staging.get("major_category") or "", staging.get("admission_type") or "", staging.get("admission_status") or "",
                    staging.get("admission_date") or "", staging.get("admission_year") or 2026, staging.get("language_requirement_type") or "",
                    staging.get("language_score_required") if staging.get("language_score_required") is not None else 0, staging.get("sat_required") if staging.get("sat_required") is not None else 0,
                    staging.get("scholarship_amount") if staging.get("scholarship_amount") is not None else 0.0,
                    staging.get("scholarship_currency") or "", staging.get("scholarship_type") or "",
                    staging.get("source_school") or "", staging.get("article_url") or "", staging.get("article_title") or "",
                    staging.get("recognition_model") or "", staging.get("data_source") or "",
                    1, "approved", comment or "", reviewer, "CURRENT_TIMESTAMP", "CURRENT_TIMESTAMP",
                    "CURRENT_TIMESTAMP", "CURRENT_TIMESTAMP", comment or ""
                ]
                exec_sql(cursor, "INSERT INTO admission_records (" + cols + ") VALUES (" + placeholders + ")", values)
                approved += 1
            except Exception:
                pass
        else:
            rejected += 1
        
        if DB_TYPE == "mysql":
            try:
                cursor.execute(
                    "INSERT INTO review_history (record_id, action, comment, reviewer) VALUES (%s, %s, %s, %s)",
                    [record_id, action, comment, reviewer])
            except Exception:
                pass
        else:
            try:
                cursor.execute(
                    "INSERT INTO review_history (record_id, action, comment, reviewer) VALUES (?, ?, ?, ?)",
                    [record_id, action, comment, reviewer])
            except Exception:
                pass
    
    conn.commit()
    conn.close()
    
    return {
        "message": f"批量审核完成：通过 {approved} 条，拒绝 {rejected} 条",
        "approved": approved,
        "rejected": rejected
    }


@app.post("/api/review/{record_id}")
async def submit_review(record_id: int, review_data: Dict[str, Any]):
    """提交单条记录审核"""
    conn = get_db_connection()
    cursor = conn.cursor()
    ph = "%s" if DB_TYPE == "mysql" else "?"
    
    exec_sql(cursor, "SELECT * FROM admission_records_staging WHERE id = " + ph, [record_id])
    row = cursor.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="记录不存在")
    
    staging = dict(row)
    
    action = review_data.get("action")  # approved / rejected
    comment = review_data.get("comment", "")
    reviewer = review_data.get("reviewer", "系统")
    
    if action not in ("approved", "rejected"):
        conn.close()
        raise HTTPException(status_code=400, detail="审核动作必须是 approved 或 rejected")
    
    exec_sql(cursor,
        "UPDATE admission_records_staging SET review_status = " + ph + ", review_comment = " + ph + ", reviewed_by = " + ph + ", reviewed_at = CURRENT_TIMESTAMP WHERE id = " + ph,
        [action, comment, reviewer, record_id])
    
    if action == "approved":
        ph = "%s" if DB_TYPE == "mysql" else "?"
        cols = (
            "student_name_cn, student_name_en, student_grade, "
            "country, country_en, university_cn, university_en, "
            "university_type, university_ranking, major_cn, major_en, "
            "major_category, admission_type, admission_status, "
            "admission_date, admission_year, language_requirement_type, "
            "language_score_required, sat_required, "
            "scholarship_amount, scholarship_currency, scholarship_type, "
            "source_school, article_url, article_title, "
            "recognition_model, data_source, "
            "status, review_status, review_note, reviewed_by, reviewed_at, promoted_at, "
            "created_at, updated_at, notes"
        )
        placeholders = (ph + ", ") * 35 + ph
        values = [
            staging.get("student_name_cn") or "", staging.get("student_name_en") or "", staging.get("student_grade") or "",
            staging.get("country") or "", staging.get("country_en") or "", staging.get("university_cn") or "", staging.get("university_en") or "",
            staging.get("university_type") or "", staging.get("university_ranking") if staging.get("university_ranking") is not None else -1, staging.get("major_cn") or "", staging.get("major_en") or "",
            staging.get("major_category") or "", staging.get("admission_type") or "", staging.get("admission_status") or "",
            staging.get("admission_date") or "", staging.get("admission_year") or 2026, staging.get("language_requirement_type") or "",
            staging.get("language_score_required") if staging.get("language_score_required") is not None else 0, staging.get("sat_required") if staging.get("sat_required") is not None else 0,
            staging.get("scholarship_amount") if staging.get("scholarship_amount") is not None else 0.0,
            staging.get("scholarship_currency") or "", staging.get("scholarship_type") or "",
            staging.get("source_school") or "", staging.get("article_url") or "", staging.get("article_title") or "",
            staging.get("recognition_model") or "", staging.get("data_source") or "",
            1, "approved", comment or "", reviewer, "CURRENT_TIMESTAMP", "CURRENT_TIMESTAMP",
            "CURRENT_TIMESTAMP", "CURRENT_TIMESTAMP", comment or ""
        ]
        exec_sql(cursor, "INSERT INTO admission_records (" + cols + ") VALUES (" + placeholders + ")", values)
    
    if DB_TYPE == "mysql":
        try:
            cursor.execute(
                "INSERT INTO review_history (record_id, action, comment, reviewer) VALUES (%s, %s, %s, %s)",
                [record_id, action, comment, reviewer])
        except Exception:
            pass
    else:
        try:
            cursor.execute(
                "INSERT INTO review_history (record_id, action, comment, reviewer) VALUES (?, ?, ?, ?)",
                [record_id, action, comment, reviewer])
        except Exception:
            pass
    
    conn.commit()
    
    action_label = "审核通过" if action == "approved" else "审核拒绝"
    write_log(operation_type="review", operation_desc=f"{action_label} 暂存记录 ID={record_id}，审核人：{reviewer}")
    
    conn.close()
    
    return {
        "message": f"审核{'通过' if action == 'approved' else '拒绝'}成功",
        "record_id": record_id,
        "action": action
    }


@app.get("/api/review/{record_id}/history")
async def get_review_history(record_id: int):
    """获取审核历史"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT id, action, comment, reviewer, created_at
        FROM review_history 
        WHERE record_id = ?
        ORDER BY created_at DESC
    """, [record_id])
    
    rows = cursor.fetchall()
    history = [dict(row) for row in rows]
    
    for h in history:
        if h.get('created_at'):
            h['created_at'] = str(h['created_at']).replace('T', ' ')[:19]
    
    conn.close()
    
    return {"history": history}


@app.get("/api/review/stats")
async def get_review_stats():
    """获取审核统计"""
    conn = get_db_connection()
    cursor = conn.cursor(pymysql.cursors.DictCursor) if DB_TYPE == "mysql" else conn.cursor()
    
    # 各状态数量 - 从暂存表查询
    cursor.execute("SELECT review_status, COUNT(*) as count FROM admission_records_staging GROUP BY review_status")
    status_counts = {row['review_status']: row['count'] for row in cursor.fetchall()}
    
    # 今日审核数量
    today_counts = {}
    try:
        if DB_TYPE == "mysql":
            cursor.execute("SELECT action, COUNT(*) as count FROM review_history WHERE DATE(created_at) = CURDATE() GROUP BY action")
        else:
            cursor.execute("SELECT action, COUNT(*) as count FROM review_history WHERE date(created_at) = date('now') GROUP BY action")
        today_counts = {row['action']: row['count'] for row in cursor.fetchall()}
    except Exception:
        today_counts = {"approved": 0, "rejected": 0}
    
    conn.close()
    
    return {
        "pending": status_counts.get('pending', 0),
        "approved": status_counts.get('approved', 0),
        "rejected": status_counts.get('rejected', 0),
        "today_approved": today_counts.get('approved', 0),
        "today_rejected": today_counts.get('rejected', 0)
    }



# 静态文件
static_path = os.path.join(os.path.dirname(__file__), "../../frontend/dist")
if os.path.exists(static_path):
    app.mount("/assets", StaticFiles(directory=static_path), name="static")

# 后端静态资源（echarts 等本地 JS 库）
backend_static = os.path.join(os.path.dirname(__file__), "../static")
if os.path.exists(backend_static):
    app.mount("/backend-static", StaticFiles(directory=backend_static), name="backend-static")


@app.post("/api/auth/login")
async def api_login(request: Request):
    """登录接口"""
    try:
        data = await request.json()
    except:
        raise HTTPException(status_code=400, detail="请求格式错误")
    
    username = data.get("username", "")
    password = data.get("password", "")
    
    if not username or not password:
        raise HTTPException(status_code=400, detail="请输入用户名和密码")
    
    user = authenticate_user(username, password)
    if not user:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    
    # 创建 access token
    access_token = create_access_token(data={
        "sub": str(user["id"]),
        "username": user["username"],
        "role": user.get("role_name", "user")
    })
    
    # 更新最后登录时间
    try:
        update_last_login(user["id"])
    except:
        pass
    
    # 记录登录日志
    write_log(user_id=user["id"], username=user["username"], operation_type="login", operation_desc=f"用户 {user['username']} 登录系统")
    
    return {
        "success": True,
        "access_token": access_token,
        "token_type": "bearer",
        "user": {
            "id": user["id"],
            "username": user["username"],
            "full_name": user.get("full_name", ""),
            "role": user.get("role_name", "user"),
            "role_id": user.get("role_id", 3),
            "role_name": user.get("role_name", "user")
        }
    }



# ==================== 系统管理 API ====================

@app.post("/api/system/logs")
async def create_operation_log(request: Request):
    """记录操作日志"""
    data = await request.json()
    conn = get_db_connection()
    cursor = conn.cursor()
    
    placeholder = "%s" if DB_TYPE == "mysql" else "?"
    exec_sql(cursor, f"INSERT INTO operation_logs (user_id, username, operation_type, operation_desc, ip_address) VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})", [data.get("user_id", 0), data.get("username", "system"), data.get("operation_type", "unknown"), data.get("operation_desc", ""), request.client.host if request.client else ""])
    conn.commit()
    conn.close()
    return {"success": True, "message": "日志已记录"}

@app.get("/api/system/logs")
async def get_operation_logs(page: int = 1, page_size: int = 20, user: str = "", type: str = ""):
    """获取操作日志列表"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    placeholder = "%s" if DB_TYPE == "mysql" else "?"
    where = []
    params = []
    if user:
        where.append(f"username LIKE {placeholder}")
        params.append(f"%{user}%")
    if type:
        where.append(f"operation_type = {placeholder}")
        params.append(type)
    
    where_sql = " WHERE " + " AND ".join(where) if where else ""
    
    # Get total count
    exec_sql(cursor, f"SELECT COUNT(*) as cnt FROM operation_logs{where_sql}", params)
    total = cursor.fetchone()["cnt"]
    
    # Get data
    offset = (page - 1) * page_size
    exec_sql(cursor, f"SELECT * FROM operation_logs{where_sql} ORDER BY created_at DESC LIMIT {placeholder} OFFSET {placeholder}", params + [page_size, offset])
    logs = [dict(row) for row in cursor.fetchall()]
    
    conn.close()
    return {"logs": logs, "total": total, "page": page, "page_size": page_size}

@app.delete("/api/system/logs/{log_id}")
async def delete_operation_log(log_id: int):
    """删除操作日志"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    placeholder = "%s" if DB_TYPE == "mysql" else "?"
    exec_sql(cursor, f"DELETE FROM operation_logs WHERE id = {placeholder}", [log_id])
    conn.commit()
    conn.close()
    return {"success": True, "message": "日志已删除"}

@app.get("/api/system/backup-settings")
async def get_backup_settings():
    """获取备份设置"""
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    exec_sql(cursor, "SELECT * FROM backup_settings LIMIT 1")
    row = cursor.fetchone()
    conn.close()
    if row:
        return dict(row)
    return {"backup_interval_days": 7, "time_span_days": 30, "auto_backup_enabled": 0}

@app.post("/api/system/backup-settings")
async def update_backup_settings(request: Request):
    """更新备份设置"""
    data = await request.json()
    conn = get_db_connection()
    cursor = conn.cursor()
    exec_sql(cursor, "UPDATE backup_settings SET backup_interval_days = ?, time_span_days = ?, auto_backup_enabled = ?, updated_at = CURRENT_TIMESTAMP", [data.get("backup_interval_days", 7), data.get("time_span_days", 30), data.get("auto_backup_enabled", 0)])
    conn.commit()
    conn.close()
    return {"success": True, "message": "备份设置已更新"}

@app.post("/api/system/backup")
async def manual_backup(request: Request):
    """手动备份操作日志"""
    import json
    import os
    from datetime import datetime, timedelta
    
    data = await request.json()
    time_span_days = data.get("time_span_days", 30)
    
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Get logs to backup
    cutoff_date = (datetime.now() - timedelta(days=time_span_days)).strftime("%Y-%m-%d %H:%M:%S")
    exec_sql(cursor, "SELECT * FROM operation_logs WHERE created_at >= ? ORDER BY created_at", [cutoff_date])
    logs = [dict(row) for row in cursor.fetchall()]
    
    if not logs:
        conn.close()
        return {"success": False, "message": "没有需要备份的日志数据"}
    
    # Create backup file
    backup_dir = os.path.join(os.path.dirname(__file__), "../backups/operation_logs")
    os.makedirs(backup_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = f"{backup_dir}/logs_backup_{timestamp}.json"
    
    with open(backup_file, "w", encoding="utf-8") as f:
        json.dump(logs, f, ensure_ascii=False, indent=2)
    
    # Update last backup time
    exec_sql(cursor, "UPDATE backup_settings SET last_backup_at = CURRENT_TIMESTAMP")
    conn.commit()
    
    # Delete backed up logs
    deleted_count = len(logs)
    exec_sql(cursor, "DELETE FROM operation_logs WHERE created_at >= ?", [cutoff_date])
    conn.commit()
    conn.close()
    
    return {
        "success": True,
        "message": "备份成功",
        "backup_file": backup_file,
        "backed_up_count": len(logs),
        "deleted_count": deleted_count
    }

@app.post("/api/system/users")
async def create_user_api(request: Request):
    """创建用户账号"""
    from app.auth import create_user, get_user_by_username
    data = await request.json()
    
    username = data.get("username", "").strip()
    password = data.get("password", "").strip()
    email = data.get("email", "").strip() or None
    full_name = data.get("full_name", "").strip() or None
    role_id = data.get("role_id", 3)
    
    if not username or not password:
        raise HTTPException(status_code=400, detail="用户名和密码不能为空")
    
    existing = get_user_by_username(username)
    if existing:
        raise HTTPException(status_code=400, detail="用户名已存在")
    
    user = create_user(username, password, email, full_name, role_id)
    
    # Log the operation
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        placeholder = "%s" if DB_TYPE == "mysql" else "?"
        exec_sql(cursor, f"INSERT INTO operation_logs (user_id, username, operation_type, operation_desc) VALUES (0, 'system', 'create_user', {placeholder})", ['创建用户: ' + username])
        conn.commit()
        conn.close()
    except Exception as e:
        pass
    
    return {"success": True, "message": "用户创建成功", "user": {"id": user["id"], "username": user["username"]}}

@app.get("/api/system/users")
async def list_users_api():
    """获取用户列表"""
    from app.auth import get_all_users, get_all_roles
    users = get_all_users()
    roles = get_all_roles()
    return {"users": users, "roles": roles}

@app.put("/api/system/users/{user_id}")
async def update_user_api(user_id: int, request: Request):
    """更新用户信息"""
    from app.auth import update_user_role, get_user_by_id, deactivate_user, activate_user, get_password_hash
    data = await request.json()
    
    action = data.get("action", "")
    
    if action == "update_role":
        role_id = data.get("role_id", 3)
        update_user_role(user_id, role_id)
        return {"success": True, "message": "角色已更新"}
    elif action == "deactivate":
        deactivate_user(user_id)
        return {"success": True, "message": "用户已停用"}
    elif action == "activate":
        activate_user(user_id)
        return {"success": True, "message": "用户已启用"}
    elif action == "reset_password":
        new_password = data.get("new_password", "")
        if not new_password:
            raise HTTPException(status_code=400, detail="请输入新密码")
        conn = get_db_connection()
        cursor = conn.cursor()
        hashed = get_password_hash(new_password)
        placeholder = "%s" if DB_TYPE == "mysql" else "?"
        exec_sql(cursor, f"UPDATE users SET hashed_password = {placeholder} WHERE id = {placeholder}", [hashed, user_id])
        conn.commit()
        conn.close()
        return {"success": True, "message": "密码已重置"}
    
    raise HTTPException(status_code=400, detail="未知操作")



@app.get("/login")
async def login_page():
    """登录页面"""
    login_path = os.path.join(os.path.dirname(__file__), "static", "login", "index.html")
    return FileResponse(login_path)


@app.get("/register")
async def register_page():
    """注册页面"""
    register_path = os.path.join(os.path.dirname(__file__), "static", "register", "index.html")
    return FileResponse(register_path)


@app.post("/api/auth/register")
async def api_register(request: Request):
    """用户注册 - 仅允许注册普通用户（role_id=3）"""
    from app.auth import create_user, get_user_by_username, get_password_hash
    import sqlite3
    
    data = await request.json()
    username = data.get("username", "").strip()
    password = data.get("password", "")
    email = data.get("email", "").strip() or None
    platform = data.get("platform", None)  # feishu / wechat / None (username+password)
    
    if not username or len(username) < 3:
        raise HTTPException(status_code=400, detail="用户名至少需要3个字符")
    
    if platform:
        # 第三方注册：platform 作为用户名前缀
        username = f"{platform}_{username}"
    elif not password or len(password) < 6:
        raise HTTPException(status_code=400, detail="密码至少需要6个字符")
    
    # 检查用户名是否已存在
    existing = get_user_by_username(username)
    if existing:
        raise HTTPException(status_code=400, detail="用户名已存在")
    
    # 注册为普通用户（role_id=3）
    try:
        user = create_user(username, password, email, username, role_id=3)
        return {
            "success": True,
            "message": "注册成功",
            "user": {
                "id": user["id"],
                "username": user["username"],
                "role_id": user["role_id"],
                "role_name": user.get("role_name", "普通用户")
            }
        }
    except HTTPException as e:
        raise e
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"注册失败：{str(e)}")


@app.get("/admin/users")
async def users_page():
    """用户管理页面"""
    users_path = os.path.join(os.path.dirname(__file__), "static", "admin", "users.html")
    return FileResponse(users_path)


# ==================== 实时监控 API ====================

@app.get("/api/monitor/status")
async def get_monitor_status():
    """获取系统实时监控状态"""
    import psutil
    from datetime import datetime, timedelta

    try:
        # 系统负载
        cpu_percent = psutil.cpu_percent(interval=0.1)
        memory = psutil.virtual_memory()
        disk = psutil.disk_usage('/')

        # 数据库状态
        conn = get_db_connection()
        cursor = conn.cursor()

        # 任务统计
        cursor.execute("""
            SELECT
                SUM(CASE WHEN task_status = 0 THEN 1 ELSE 0 END) as pending,
                SUM(CASE WHEN task_status = 1 THEN 1 ELSE 0 END) as retrying,
                SUM(CASE WHEN task_status = 2 THEN 1 ELSE 0 END) as processing,
                SUM(CASE WHEN task_status = 3 THEN 1 ELSE 0 END) as completed,
                SUM(CASE WHEN task_status = 4 THEN 1 ELSE 0 END) as failed,
                SUM(CASE WHEN task_status = 5 THEN 1 ELSE 0 END) as cancelled,
                COUNT(*) as total
            FROM collection_tasks
        """)
        task_stats = dict(cursor.fetchone())

        # 最近1小时完成速率
        one_hour_ago = (datetime.now() - timedelta(hours=1)).strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute("""
            SELECT COUNT(*), AVG(extracted_count)
            FROM collection_tasks
            WHERE task_status = 3 AND completed_at >= ?
        """, [one_hour_ago])
        row = cursor.fetchone()
        recent_completed = row[0] or 0
        avg_extracted = row[1] or 0

        # 并发设置
        exec_sql(cursor, "SELECT enabled, max_concurrent FROM concurrency_settings WHERE id = 1")
        concurrency_row = cursor.fetchone()
        concurrency_enabled = bool(concurrency_row[0]) if concurrency_row else False
        max_concurrent = concurrency_row[1] if concurrency_row else 2

        conn.close()

        return {
            "system": {
                "cpu_percent": round(cpu_percent, 1),
                "memory_percent": round(memory.percent, 1),
                "memory_used_mb": round(memory.used / 1024 / 1024, 0),
                "memory_total_mb": round(memory.total / 1024 / 1024, 0),
                "disk_percent": round(disk.percent, 1),
                "timestamp": datetime.now().isoformat()
            },
            "tasks": {
                "pending": task_stats.get('pending', 0),
                "retrying": task_stats.get('retrying', 0),
                "processing": task_stats.get('processing', 0),
                "completed": task_stats.get('completed', 0),
                "failed": task_stats.get('failed', 0),
                "cancelled": task_stats.get('cancelled', 0),
                "total": task_stats.get('total', 0)
            },
            "performance": {
                "recent_completed_1h": recent_completed,
                "avg_extracted_per_task": round(avg_extracted, 2),
                "rate_per_hour": recent_completed,
                "concurrency_enabled": concurrency_enabled,
                "max_concurrent": max_concurrent
            }
        }
    except Exception as e:
        return {
            "error": str(e),
            "timestamp": datetime.now().isoformat()
        }


@app.get("/{full_path:path}")
async def serve_spa(full_path: str):
    if full_path.startswith("api/") or full_path.startswith("ws/"):
        raise HTTPException(status_code=404, detail="Not Found")
    static_file = os.path.join(static_path, full_path)
    if os.path.exists(static_file) and os.path.isfile(static_file):
        # JS/CSS 带 hash 的文件可以长期缓存，HTML 必须无缓存
        if full_path.endswith(('.js', '.css')):
            return FileResponse(static_file, headers={"Cache-Control": "public, max-age=31536000, immutable"})
        return FileResponse(static_file)
    index_path = os.path.join(static_path, "index.html")
    if os.path.exists(index_path):
        # HTML 文件必须无缓存，确保浏览器加载最新版本
        return FileResponse(index_path, headers={"Cache-Control": "no-cache, no-store, must-revalidate", "Pragma": "no-cache", "Expires": "0"})
    return JSONResponse({"message": "大学录取信息整理系统 API", "version": "1.0.0", "docs": "/docs"})



